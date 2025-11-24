from collections import defaultdict
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.db.models.functions import ExtractMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.utils.timezone import make_aware

from main.calculations import _month_bounds
from main.models import (
    AccountEntry,
    BillingAccount,
    Order,
    PaymentAttempt,
    StarlinkKitInventory,
    Subscription,
    Ticket,
    User,
    WalletTransaction,
)
from user.permissions import require_staff_role


# Create your views here.
@login_required(login_url="login_page")
@require_staff_role(["admin", "manager", "finance"])
def dashboard_bi(request):
    template = "dashboard_bi_page.html"
    return render(request, template)


@login_required(login_url="login_page")
@require_staff_role(["admin", "manager", "finance"])
def kpis_monthly(request):
    """
    Return JSON with:
      - monthly_revenue (USD)
      - active_subscriptions (count)
      - arpu (USD)
      - churn_rate (percent string, e.g., "3.2%")
    Period = current calendar month in user's timezone.
    """
    now = timezone.localtime()
    month_start, month_end = _month_bounds(now)

    # ----- Monthly Revenue -----
    # Use PaymentAttempt status "completed" within the month (transaction_time preferred)
    paid_filter = Q(status="completed") & (
        Q(transaction_time__gte=month_start, transaction_time__lt=month_end)
        | (
            Q(transaction_time__isnull=True)
            & Q(created_at__gte=month_start, created_at__lt=month_end)
        )
    )
    monthly_revenue = PaymentAttempt.objects.filter(paid_filter).aggregate(
        total=Sum("amount")
    ).get("total") or Decimal("0.00")

    # ----- Active Subscriptions (in period) -----
    # Any subscription that overlaps the month window:
    # started_at <= month_end and (ended_at is null or ended_at >= month_start)
    active_overlap_q = Subscription.objects.filter(
        Q(started_at__lte=month_end.date())
        & (Q(ended_at__isnull=True) | Q(ended_at__gte=month_start.date()))
    )
    active_subscriptions = active_overlap_q.count()

    # Distinct active users (for ARPU denominator)
    active_users_in_month = active_overlap_q.values("user_id").distinct().count()

    # ----- ARPU -----
    arpu = Decimal("0.00")
    if active_users_in_month > 0:
        arpu = (Decimal(monthly_revenue) / Decimal(active_users_in_month)).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

    # ----- Churn Rate -----
    # Approximation:
    #   churned_this_month = subs that ended within [month_start, month_end)
    #   opening_active = subs active at the START of the month
    churned_this_month = Subscription.objects.filter(
        status="cancelled",
        ended_at__gte=month_start.date(),
        ended_at__lt=month_end.date(),
    ).count()

    opening_active = Subscription.objects.filter(
        Q(started_at__lte=month_start.date())
        & (Q(ended_at__isnull=True) | Q(ended_at__gte=month_start.date()))
    ).count()

    if opening_active > 0:
        churn_rate_val = (
            Decimal(churned_this_month) / Decimal(opening_active)
        ) * Decimal(100)
    else:
        churn_rate_val = Decimal("0.00")

    # Format outputs
    out = {
        "monthly_revenue": str(
            Decimal(monthly_revenue).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        ),
        "active_subscriptions": active_subscriptions,
        "arpu": str(arpu),  # USD
        "churn_rate": f"{churn_rate_val.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}%",
    }
    return JsonResponse(out, status=200)


@login_required(login_url="login_page")
@require_staff_role(["admin", "manager", "finance"])
def kpis_annual(request):
    """
    Return current-year monthly revenue and subscription counts (new subs started).
    Response:
      {
        "year": 2025,
        "months": ["Jan", ... , "Dec"],
        "revenue_by_month": [0, 1234.56, ...],
        "subs_started_by_month": [0, 3, ...]
      }
    """
    now = timezone.now()
    year = now.year

    # --- Revenue: sum of PAID orders' total_price by month (current year)
    paid_orders = (
        Order.objects.filter(payment_status="paid", created_at__year=year)
        .annotate(m=ExtractMonth("created_at"))
        .values("m")
        .annotate(total=Sum("total_price"))
        .values("m", "total")
    )
    revenue_map = {row["m"]: float(row["total"] or 0.0) for row in paid_orders}

    # --- Subscriptions: count subscriptions started per month (current year)
    subs_started = (
        Subscription.objects.filter(started_at__year=year)
        .annotate(m=ExtractMonth("started_at"))
        .values("m")
        .annotate(cnt=Count("id"))
        .values("m", "cnt")
    )
    subs_map = {row["m"]: int(row["cnt"] or 0) for row in subs_started}

    # Build full 12-month arrays
    month_labels = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]
    revenue = [round(revenue_map.get(i, 0.0), 2) for i in range(1, 13)]
    subs = [subs_map.get(i, 0) for i in range(1, 13)]

    return JsonResponse(
        {
            "year": year,
            "months": month_labels,
            "revenue_by_month": revenue,
            "subs_started_by_month": subs,
        }
    )


def _parse_date(s):
    if not s:
        return None
    try:
        # naive -> aware (server timezone)
        return make_aware(datetime.strptime(s, "%Y-%m-%d"))
    except Exception:
        return None


def _is_staff(user):
    return user.is_authenticated and user.is_staff


# ---------- XLSX utility ----------


def _xlsx_http_response(filename: str, headers: list[str], rows_iterable):
    """
    Build an XLSX file in-memory and return an HttpResponse.
    - headers: list of column titles
    - rows_iterable: iterable of lists/tuples (each is a row)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Write header
    ws.append(headers)

    # Write rows, track max width per column for autosize
    max_widths = [len(str(h)) for h in headers]

    for row in rows_iterable:
        ws.append(row)
        for col_idx, cell_val in enumerate(row, start=1):
            length = len(str(cell_val)) if cell_val is not None else 0
            if length > max_widths[col_idx - 1]:
                max_widths[col_idx - 1] = length

    # Auto-size columns (with some padding and a sane max)
    for i, w in enumerate(max_widths, start=1):
        adj = min(w + 2, 60)
        ws.column_dimensions[get_column_letter(i)].width = adj

    # Save to bytes
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ---------- EXPORTS (XLSX) ----------


@login_required(login_url="login_page")
@require_staff_role(["admin", "manager", "finance"])
def export_customers_csv(request):  # endpoint name preserved
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()  # Optional: Active/Inactive

    qs = (
        User.objects.filter(is_staff=False)
        .order_by("-id_user")
        .only("id_user", "full_name", "email", "phone", "is_active", "is_verified")
    )
    if q:
        qs = qs.filter(
            Q(full_name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q)
        )
    if status in ("Active", "Inactive"):
        qs = qs.filter(is_active=(status == "Active"))

    headers = [
        "ID",
        "Full Name",
        "Email",
        "Phone",
        "Account Active",
        "Verified",
        "KYC Status",
    ]

    def rows():
        for u in qs.iterator():
            yield [
                u.id_user,
                u.full_name,
                u.email,
                u.phone,
                "Yes" if u.is_active else "No",
                "Yes" if u.is_verified else "No",
                u.get_kyc_status(),
            ]

    return _xlsx_http_response("customers.xlsx", headers, rows())


@login_required(login_url="login_page")
@require_staff_role(["admin", "manager", "finance"])
def export_orders_csv(request):  # endpoint name preserved
    start = _parse_date(request.GET.get("start"))
    end = _parse_date(request.GET.get("end"))

    qs = Order.objects.select_related("user", "plan").order_by("-created_at")
    if start:
        qs = qs.filter(created_at__gte=start)
    if end:
        qs = qs.filter(created_at__lt=end)

    headers = [
        "Order Ref",
        "Customer",
        "Email",
        "Plan",
        "Total (USD)",
        "Payment Status",
        "Status",
        "Created At",
        "Installed",
        "Installation Date",
    ]

    def rows():
        for o in qs.iterator():
            yield [
                o.order_reference or o.id,
                getattr(o.user, "full_name", "") or "",
                getattr(o.user, "email", "") or "",
                getattr(o.plan, "name", "") or "",
                f"{o.total_price or 0:.2f}",
                o.payment_status,
                o.status,
                o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "",
                "Yes" if o.is_installed else "No",
                o.installation_date.strftime("%Y-%m-%d %H:%M")
                if o.installation_date
                else "",
            ]

    return _xlsx_http_response("orders.xlsx", headers, rows())


@login_required(login_url="login_page")
@require_staff_role(["admin", "manager", "finance"])
def export_subscriptions_csv(request):  # endpoint name preserved
    start = _parse_date(request.GET.get("start"))
    end = _parse_date(request.GET.get("end"))
    status = (request.GET.get("status") or "").strip()  # active/suspended/cancelled

    qs = Subscription.objects.select_related("user", "plan").order_by("-started_at")
    if start:
        qs = qs.filter(
            Q(started_at__gte=start.date()) | Q(next_billing_date__gte=start.date())
        )
    if end:
        qs = qs.filter(
            Q(started_at__lt=end.date()) | Q(next_billing_date__lt=end.date())
        )
    if status in ("active", "suspended", "cancelled"):
        qs = qs.filter(status=status)

    headers = [
        "Subscription ID",
        "Customer",
        "Email",
        "Plan",
        "Billing Cycle",
        "Status",
        "Start Date",
        "Next Billing",
        "End Date",
    ]

    def rows():
        for s in qs.iterator():
            yield [
                s.id,
                getattr(s.user, "full_name", "") or "",
                getattr(s.user, "email", "") or "",
                getattr(s.plan, "name", "") or "",
                s.billing_cycle,
                s.status,
                s.started_at.isoformat() if s.started_at else "",
                s.next_billing_date.isoformat() if s.next_billing_date else "",
                s.ended_at.isoformat() if s.ended_at else "",
            ]

    return _xlsx_http_response("subscriptions.xlsx", headers, rows())


@login_required(login_url="login_page")
@require_staff_role(["admin", "manager", "finance"])
def export_revenue_monthly_csv(request):  # endpoint name preserved
    """
    Sums AccountEntry by month for:
      - invoices (entry_type='invoice') as positive charges
      - payments (entry_type='payment') as credits (-)
    Optional ?start=YYYY-MM-DD&end=YYYY-MM-DD
    """
    start = _parse_date(request.GET.get("start"))
    end = _parse_date(request.GET.get("end"))

    entries = AccountEntry.objects.all()
    if start:
        entries = entries.filter(created_at__gte=start)
    if end:
        entries = entries.filter(created_at__lt=end)

    from django.db.models.functions import TruncMonth

    monthly = (
        entries.annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(
            invoices=Sum("amount_usd", filter=Q(entry_type="invoice")),
            payments=Sum("amount_usd", filter=Q(entry_type="payment")),
            count_invoices=Count("id", filter=Q(entry_type="invoice")),
            count_payments=Count("id", filter=Q(entry_type="payment")),
        )
        .order_by("month")
    )

    headers = [
        "Month",
        "Invoice Amount (USD)",
        "Payments (USD)",
        "#Invoices",
        "#Payments",
        "Net (USD)",
    ]

    def rows():
        for row in monthly:
            inv = row["invoices"] or 0
            pay = row["payments"] or 0
            net = (inv or 0) + (pay or 0)
            label = row["month"].strftime("%Y-%m")
            yield [
                label,
                f"{inv:.2f}",
                f"{pay:.2f}",
                row["count_invoices"],
                row["count_payments"],
                f"{net:.2f}",
            ]

    return _xlsx_http_response("revenue_monthly.xlsx", headers, rows())


@login_required(login_url="login_page")
@require_staff_role(["admin", "manager", "finance"])
def export_tickets_csv(request):  # endpoint name preserved
    start = _parse_date(request.GET.get("start"))
    end = _parse_date(request.GET.get("end"))
    status = (request.GET.get("status") or "").strip().lower()

    qs = Ticket.objects.select_related("user").order_by("-created_at")
    if start:
        qs = qs.filter(created_at__gte=start)
    if end:
        qs = qs.filter(created_at__lt=end)
    if status in ("open", "pending", "closed"):
        qs = qs.filter(status=status)

    headers = [
        "Ticket #",
        "Subject",
        "Category",
        "Priority",
        "Status",
        "Customer",
        "Email",
        "Created At",
        "Updated At",
    ]

    def rows():
        for t in qs.iterator():
            yield [
                t.pk,
                t.subject,
                t.category,
                t.priority,
                t.status,
                getattr(t.user, "full_name", "") or "",
                getattr(t.user, "email", "") or "",
                t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
                t.updated_at.strftime("%Y-%m-%d %H:%M") if t.updated_at else "",
            ]

    return _xlsx_http_response("tickets.xlsx", headers, rows())


@login_required(login_url="login_page")
@require_staff_role(["admin", "manager", "finance"])
def export_inventory_csv(request):  # endpoint name preserved
    qs = StarlinkKitInventory.objects.select_related("kit").order_by("kit_number")

    headers = [
        "Kit Number",
        "Serial Number",
        "Model",
        "Kit Name",
        "Kit Type",
        "Assigned",
        "Assigned Order Ref",
    ]

    def rows():
        for it in qs.iterator():
            yield [
                it.kit_number or "",
                it.serial_number or "",
                it.model or "",
                getattr(it.kit, "name", "") or "",
                getattr(it.kit, "kit_type", "") or "",
                "Yes" if it.is_assigned else "No",
                getattr(getattr(it, "assigned_to_order", None), "order_reference", "")
                or "",
            ]

    return _xlsx_http_response("inventory.xlsx", headers, rows())


# ─────────────────────────────────────────────────────────────────────────────
# GENERAL LEDGER (AccountEntry ± WalletTransaction)  → XLSX
# ─────────────────────────────────────────────────────────────────────────────


@login_required(login_url="login_page")
@require_staff_role(["admin", "manager", "finance"])
def export_general_ledger_csv(request):  # URL name mirrors your pattern, returns XLSX
    """
    General Ledger export (XLSX) with running balance per BillingAccount.

    Query params:
      - start=YYYY-MM-DD
      - end=YYYY-MM-DD
      - entry_type=invoice|payment|credit_note|adjustment|tax
      - account_id=<int> (BillingAccount.pk)
      - q=<search in user full_name/email>
      - include_wallet=1  (merge Wallet transactions into a separate section)
    """
    start = _parse_date(request.GET.get("start"))
    end = _parse_date(request.GET.get("end"))
    entry_type = (request.GET.get("entry_type") or "").strip()
    include_wallet = (request.GET.get("include_wallet") or "").strip() in (
        "1",
        "true",
        "yes",
    )
    account_id = request.GET.get("account_id")
    q = (request.GET.get("q") or "").strip()

    # Base queryset for AccountEntry with traceability
    entries = AccountEntry.objects.select_related(
        "account__user", "order", "subscription", "payment"
    ).order_by("account_id", "created_at", "id")

    if start:
        entries = entries.filter(created_at__gte=start)
    if end:
        entries = entries.filter(created_at__lt=end)
    if entry_type in {"invoice", "payment", "credit_note", "adjustment", "tax"}:
        entries = entries.filter(entry_type=entry_type)
    if account_id:
        entries = entries.filter(account_id=account_id)
    if q:
        entries = entries.filter(
            Q(account__user__full_name__icontains=q)
            | Q(account__user__email__icontains=q)
        )

    # Optionally prefetch wallet txns for same filtered users
    wallet_txns = None
    if include_wallet:
        wallet_txns = WalletTransaction.objects.select_related(
            "wallet__user", "order", "payment_attempt"
        ).order_by("wallet_id", "created_at", "id")

        # Limit wallet transactions to overlapping user set if q/account filters present
        if account_id:
            # Constrain to the account's user only
            try:
                acc_user = (
                    BillingAccount.objects.select_related("user")
                    .get(pk=account_id)
                    .user
                )
                wallet_txns = wallet_txns.filter(wallet__user=acc_user)
            except BillingAccount.DoesNotExist:
                wallet_txns = wallet_txns.none()

        if q:
            wallet_txns = wallet_txns.filter(
                Q(wallet__user__full_name__icontains=q)
                | Q(wallet__user__email__icontains=q)
            )
        if start:
            wallet_txns = wallet_txns.filter(created_at__gte=start)
        if end:
            wallet_txns = wallet_txns.filter(created_at__lt=end)

    # ── Build rows with running balance per account ──────────────────────────
    headers = [
        "Date",
        "Account ID",
        "Customer",
        "Entry Type",
        "Source",
        "Description / Note",
        "Order Ref",
        "Subscription ID",
        "Payment Ref",
        "Period Start",
        "Period End",
        "External Ref",
        "Debit (USD)",
        "Credit (USD)",
        "Running Balance (USD)",  # per Account (or per Wallet when in the wallet section)
    ]

    def format_customer(u):
        if not u:
            return ""
        return u.full_name or u.email or u.username

    def as_dc(amount: Decimal):
        """Split signed amount into (debit, credit) with positive numbers."""
        if amount is None:
            return ("", "")
        amt = Decimal(amount)
        if amt >= 0:
            return (f"{amt:.2f}", "")
        return ("", f"{(-amt):.2f}")

    # Running balances keyed by account id
    running = defaultdict(lambda: Decimal("0.00"))

    def account_rows():
        for e in entries.iterator():
            acc_id = e.account_id
            # Update running balance (AccountEntry rule: +ve = debit/charge; -ve = credit/payment)
            running[acc_id] += e.amount_usd or Decimal("0.00")

            debit, credit = as_dc(e.amount_usd or Decimal("0.00"))
            user = getattr(e.account, "user", None)

            yield [
                e.created_at.strftime("%Y-%m-%d %H:%M") if e.created_at else "",
                acc_id or "",
                format_customer(user),
                e.entry_type or "",
                "ACCOUNT",
                e.description or "",
                (getattr(e.order, "order_reference", None) or (e.order_id or "")) or "",
                e.subscription_id or "",
                getattr(e.payment, "order_number", "")
                or getattr(e.payment, "provider_reference", "")
                or "",
                e.period_start.isoformat() if e.period_start else "",
                e.period_end.isoformat() if e.period_end else "",
                e.external_ref or "",
                debit,
                credit,
                f"{running[acc_id]:.2f}",
            ]

    # Wallet section (optional): running balance per wallet id (separate from account)
    wallet_running = defaultdict(lambda: Decimal("0.00"))

    def wallet_rows():
        if wallet_txns is None:
            return
        for t in wallet_txns.iterator():
            wid = t.wallet_id
            # WalletTransaction amounts are always positive; direction in tx_type
            # CREDIT increases wallet balance; DEBIT decreases
            amt = Decimal(t.amount or 0)
            if t.tx_type == WalletTransaction.Type.CREDIT:
                wallet_running[wid] += amt
                debit, credit = (
                    "",
                    f"{0:.2f}",
                )  # For GL readability we keep wallet semantics:
                # present credits as positive credit column? or as debit?
                # We'll represent wallet inflow as CREDIT column (right side), outflow as DEBIT.
                debit = ""
                credit = f"{amt:.2f}"
            else:  # DEBIT
                wallet_running[wid] -= amt
                debit = f"{amt:.2f}"
                credit = ""

            user = getattr(t.wallet, "user", None)
            order_ref = getattr(t.order, "order_reference", None) or (t.order_id or "")
            payref = (
                getattr(t.payment_attempt, "order_number", "")
                or getattr(t.payment_attempt, "provider_reference", "")
                or ""
            )

            yield [
                t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
                f"W{wid}",
                format_customer(user),
                t.tx_type.upper(),
                "WALLET",
                t.note or "",
                order_ref or "",
                "",  # no subscription link on wallet ledger
                payref,
                "",
                "",
                "",  # no period/external refs
                debit,
                credit,
                f"{wallet_running[wid]:.2f}",
            ]

    # Compose all rows: first AccountEntry, then (optionally) Wallet section with a blank separator
    def rows():
        # Account ledger
        for r in account_rows():
            yield r
        # Wallet ledger
        if include_wallet:
            # spacer row
            yield [
                "",
                "",
                "",
                "",
                "",
                "— WALLET TRANSACTIONS —",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
            for r in wallet_rows():
                yield r

    return _xlsx_http_response("general_ledger.xlsx", headers, rows())


@login_required(login_url="login_page")
@require_staff_role(["admin", "manager", "finance"])
def export_trial_balance_csv(request):
    """
    Trial Balance: sums AccountEntry over the period by BillingAccount.
    Displays total Debits, Credits, and Net (Debit - Credit).
    """
    start = _parse_date(request.GET.get("start"))
    end = _parse_date(request.GET.get("end"))
    q = (request.GET.get("q") or "").strip()

    qs = AccountEntry.objects.select_related("account__user")
    if start:
        qs = qs.filter(created_at__gte=start)
    if end:
        qs = qs.filter(created_at__lt=end)
    if q:
        qs = qs.filter(
            Q(account__user__full_name__icontains=q)
            | Q(account__user__email__icontains=q)
        )

    # Aggregate per account
    per_account = defaultdict(
        lambda: {"debit": Decimal("0.00"), "credit": Decimal("0.00"), "user": None}
    )
    for e in qs.order_by("account_id").iterator():
        user = getattr(e.account, "user", None)
        amt = Decimal(e.amount_usd or 0)
        if amt >= 0:
            per_account[e.account_id]["debit"] += amt
        else:
            per_account[e.account_id]["credit"] += -amt
        per_account[e.account_id]["user"] = user

    headers = [
        "Account ID",
        "Customer",
        "Total Debits (USD)",
        "Total Credits (USD)",
        "Net (USD)",
    ]

    def rows():
        for acc_id, v in sorted(per_account.items(), key=lambda x: x[0] or 0):
            user = v["user"]
            total_debit = v["debit"]
            total_credit = v["credit"]
            net = total_debit - total_credit
            name = (
                user.full_name
                if user and user.full_name
                else (user.email if user else "")
            ) or ""
            yield [
                acc_id,
                name,
                f"{total_debit:.2f}",
                f"{total_credit:.2f}",
                f"{net:.2f}",
            ]

    return _xlsx_http_response("trial_balance.xlsx", headers, rows())
