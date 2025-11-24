import os
import re
from datetime import datetime, timedelta
from datetime import timezone as dt_tz
from decimal import Decimal
from io import BytesIO
from urllib.parse import unquote

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from xhtml2pdf import pisa

from django.conf import settings as django_settings
from django.contrib.auth.decorators import login_required
from django.contrib.staticfiles import finders
from django.db import models
from django.db.models import Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST

from billing_management.billing_helpers import _decode_cursor, _encode_cursor
from billing_management.services.invoice_grouping import group_invoice_lines_by_order
from main.models import (
    AccountEntry,
    CompanySettings,
    ConsolidatedInvoice,
    FxRate,
    Invoice,
    Order,
    User,
)
from main.services.region_resolver import resolve_region_from_coords
from user.permissions import require_staff_role


def resolve_uri(uri, rel=None):
    """Convert HTML URIs (media/static) to absolute system paths for xhtml2pdf."""
    # If it's already an absolute file path, return as is
    if uri.startswith("file://"):
        return uri.replace("file://", "")
    # Media
    media_url = getattr(django_settings, "MEDIA_URL", "/media/")
    media_root = getattr(django_settings, "MEDIA_ROOT", "")
    if media_url and uri.startswith(media_url):
        path = os.path.join(media_root, uri.replace(media_url, ""))
        return path
    # Static
    static_url = getattr(django_settings, "STATIC_URL", "/static/")
    static_root = getattr(django_settings, "STATIC_ROOT", "")
    if static_url and uri.startswith(static_url):
        # Prefer STATIC_ROOT, fallback to finders
        if static_root:
            path = os.path.join(static_root, uri.replace(static_url, ""))
            if os.path.exists(path):
                return path
        found = finders.find(uri.replace(static_url, ""))
        if found:
            return found
    # Absolute URL (http/https) — xhtml2pdf may not fetch; return as-is
    return uri


def _money_format(val):
    try:
        return f"${Decimal(val or 0).quantize(Decimal('0.01')):,.2f}"
    except Exception:
        return "$0.00"


def _build_invoice_context(inv, cs):
    # Compute taxes similar to invoice_json_by_number
    subtotal = Decimal(inv.subtotal or 0)
    stored_vat = Decimal(getattr(inv, "vat_amount", 0) or 0)
    stored_exc = Decimal(getattr(inv, "excise_amount", 0) or 0)
    vat_guess = Decimal("0.00")
    exc_guess = Decimal("0.00")

    # Determine CDF equivalent if configured and invoice is in USD
    fx_cdf_amount = None
    fx_rate_used = None
    try:
        if (
            getattr(cs, "show_prices_in_cdf", False) or (inv.currency or "USD") == "USD"
        ) and inv.issued_at:
            fx_rate_used = FxRate.get_rate(inv.issued_at.date(), pair="USD/CDF")
            if fx_rate_used:
                # Use grand_total if available; otherwise recomputed below after we have tax_total
                pass
    except Exception:
        fx_rate_used = None

    orders = []
    try:
        orders = [ol.order for ol in inv.order_links.select_related("order")]
        if not orders:
            ords = []
            for ln in inv.lines.all():
                if getattr(ln, "order", None) and ln.order not in ords:
                    ords.append(ln.order)
            orders = ords
    except Exception:
        orders = []

    if (stored_vat + stored_exc) > 0 or Decimal(inv.tax_total or 0) == 0:
        vat_guess = stored_vat
        exc_guess = stored_exc
        if orders and exc_guess == 0:
            exc_from_orders = Decimal("0.00")
            for o in orders:
                taxes_qs = o.taxes.all()
                exc_from_orders += sum(
                    (
                        Decimal(t.amount)
                        for t in taxes_qs
                        if getattr(t, "kind", "").upper() in ("EXCISE", "EXC")
                    ),
                    Decimal("0.00"),
                )
            if exc_from_orders > 0:
                exc_guess = exc_from_orders
    else:
        if orders:
            for o in orders:
                taxes_qs = o.taxes.all()
                vat_guess += sum(
                    (
                        Decimal(t.amount)
                        for t in taxes_qs
                        if getattr(t, "kind", "").upper() == "VAT"
                    ),
                    Decimal("0.00"),
                )
                exc_guess += sum(
                    (
                        Decimal(t.amount)
                        for t in taxes_qs
                        if getattr(t, "kind", "").upper() in ("EXCISE", "EXC")
                    ),
                    Decimal("0.00"),
                )
        else:
            try:
                vat_rate = Decimal(
                    inv.vat_rate_percent or cs.vat_rate_percent or 0
                ) / Decimal("100")
            except Exception:
                vat_rate = Decimal("0")
            try:
                exc_rate = (
                    Decimal(inv.excise_rate_percent or cs.excise_rate_percent or 0)
                    / Decimal("100")
                    if (inv.excise_rate_percent or cs.excise_rate_percent)
                    else Decimal("0")
                )
            except Exception:
                exc_rate = Decimal("0")
            vat_guess = (
                (subtotal * vat_rate).quantize(Decimal("0.01"))
                if vat_rate
                else Decimal("0.00")
            )
            exc_guess = (
                (subtotal * exc_rate).quantize(Decimal("0.01"))
                if exc_rate
                else Decimal("0.00")
            )

    tax_total = Decimal(inv.tax_total or (vat_guess + exc_guess))
    grand_total = Decimal(inv.grand_total or (subtotal + tax_total))

    # Compute CDF equivalent if we have a rate
    if fx_rate_used:
        try:
            fx_cdf_amount = (grand_total * Decimal(fx_rate_used)).quantize(
                Decimal("0.01")
            )
        except Exception:
            fx_cdf_amount = None

    # Items
    items = []
    for idx, ln in enumerate(
        inv.lines.exclude(kind__iexact="tax")
        .exclude(kind__iexact="tax_adjust")
        .order_by("id"),
        start=1,
    ):
        items.append(
            {
                "idx": idx,
                "description": ln.description or ln.kind or "Item",
                "quantity": ln.quantity,
                "unit_price": _money_format(ln.unit_price),
                "line_total": _money_format(ln.line_total),
            }
        )

    # Company
    address = ", ".join(
        filter(None, [cs.street_address, cs.city, cs.province, cs.country])
    )
    logo_url = None
    try:
        if getattr(cs, "logo", None) and hasattr(cs.logo, "url") and cs.logo.url:
            # Only include logo if it resolves to a local file path that exists (xhtml2pdf limitation)
            candidate = cs.logo.url
            resolved = resolve_uri(candidate)
            if resolved and os.path.exists(resolved):
                logo_url = candidate
    except Exception:
        logo_url = None

    company = {
        "legal_name": cs.legal_name or cs.trade_name,
        "trade_name": cs.trade_name or cs.legal_name,
        "email": cs.email,
        "phone": cs.phone,
        "website": cs.website,
        "address": address,
        "street_address": getattr(cs, "street_address", ""),
        "city": getattr(cs, "city", ""),
        "province": getattr(cs, "province", ""),
        "country": getattr(cs, "country", ""),
        "rccm": cs.rccm,
        "id_nat": cs.id_nat,
        "nif": cs.nif,
        "arptc_license": getattr(cs, "arptc_license", ""),
        "tax_regime": getattr(cs, "tax_regime", ""),
        "tax_regime_label": (
            cs.get_tax_regime_display()
            if hasattr(cs, "get_tax_regime_display")
            else getattr(cs, "tax_regime", "")
        ),
        "logo_url": logo_url,
        "currency": inv.currency or cs.default_currency or "USD",
        "footer_text": cs.footer_text_en or cs.footer_text_fr,
        "footer_text_fr": getattr(cs, "footer_text_fr", ""),
        "footer_text_en": getattr(cs, "footer_text_en", ""),
        # Banking & Mobile Money
        "bank_name": getattr(cs, "bank_name", ""),
        "bank_account_name": getattr(cs, "bank_account_name", ""),
        "bank_account_number_usd": getattr(cs, "bank_account_number_usd", ""),
        "bank_account_number_cdf": getattr(cs, "bank_account_number_cdf", ""),
        "bank_swift": getattr(cs, "bank_swift", ""),
        "bank_branch": getattr(cs, "bank_branch", ""),
        "bank_iban": getattr(cs, "bank_iban", ""),
        "mm_provider": getattr(cs, "mm_provider", ""),
        "mm_number": getattr(cs, "mm_number", ""),
        "payment_instructions": getattr(cs, "payment_instructions", ""),
        # Branding & Signature
        "signatory_name": getattr(cs, "signatory_name", ""),
        "signatory_title": getattr(cs, "signatory_title", ""),
        "stamp": getattr(cs, "stamp", None),
        "signature": getattr(cs, "signature", None),
        # Compliance & Legal
        "tax_office_name": getattr(cs, "tax_office_name", ""),
        "legal_notes": getattr(cs, "legal_notes", ""),
        # Payment terms
        "payment_terms_days": getattr(cs, "payment_terms_days", 0),
    }

    # Bill to
    bill_to_name = inv.bill_to_name or (
        getattr(inv.user, "full_name", None) or getattr(inv.user, "username", "")
    )

    # Build bilingual labels inline (simple EN/FR) without relying on translation system in PDF
    bilingual_notes = [
        "Prices in USD unless otherwise stated. Official exchange rate applied when displaying CDF. / Prix en USD sauf indication contraire. Taux de change officiel appliqué pour l’affichage en CDF.",
        "Taxes per DRC rules. VAT base may include excise where applicable. / Taxes selon la réglementation RDC. La base TVA peut inclure l’accise, le cas échéant.",
    ]
    if fx_rate_used:
        bilingual_notes.insert(
            0,
            f"FX rate used (USD→CDF): {fx_rate_used} / Taux appliqué (USD→CDF) : {fx_rate_used}",
        )

    # Get tax rate percentages for display in template
    # Use the rates snapshotted on the invoice, or fall back to CompanySettings defaults
    vat_rate_percent = None
    excise_rate_percent = None
    try:
        vat_rate_percent = Decimal(inv.vat_rate_percent or cs.vat_rate_percent or 0)
    except Exception:
        vat_rate_percent = Decimal("0")
    try:
        excise_rate_percent = (
            Decimal(inv.excise_rate_percent or cs.excise_rate_percent or 0)
            if (inv.excise_rate_percent or cs.excise_rate_percent)
            else Decimal("0")
        )
    except Exception:
        excise_rate_percent = Decimal("0")

    # Group invoice lines by order for display
    order_grouping = group_invoice_lines_by_order(inv)

    # Prepare header-friendly date formats (ISO like 2025-11-12) with fallback due date
    issued_dt_local = timezone.localtime(inv.issued_at) if inv.issued_at else None
    due_dt_local = (
        timezone.localtime(inv.due_at)
        if getattr(inv, "due_at", None)
        else (
            issued_dt_local
            + timedelta(days=int(getattr(cs, "payment_terms_days", 0) or 0))
            if issued_dt_local
            else None
        )
    )
    issued_iso = issued_dt_local.strftime("%Y-%m-%d") if issued_dt_local else "—"
    due_iso = due_dt_local.strftime("%Y-%m-%d") if due_dt_local else "—"

    context = {
        "company": company,
        "invoice": {
            "number": inv.number,
            "issued_at": timezone.localtime(inv.issued_at).strftime("%d %b %Y")
            if inv.issued_at
            else "—",
            "due_at": timezone.localtime(inv.due_at).strftime("%d %b %Y")
            if getattr(inv, "due_at", None)
            else "—",
            # Header-friendly ISO dates and computed due date fallback
            "issued_at_iso": issued_iso,
            "due_at_iso": due_iso,
            "po_ref": getattr(inv, "po_reference", "") or "",
            "status": getattr(inv, "status", "").upper(),
            "currency": inv.currency or company["currency"],
            "subtotal": _money_format(subtotal),
            "excise": _money_format(exc_guess),
            "vat": _money_format(vat_guess),
            "tax_total": _money_format(tax_total),
            "grand_total": _money_format(grand_total),
            "cdf_amount": (
                f"CDF {fx_cdf_amount:,.2f}" if fx_cdf_amount is not None else None
            ),
            "bill_to_name": bill_to_name,
            "bill_to_email": getattr(inv, "bill_to_email", "")
            or getattr(inv.user, "email", ""),
            "bill_to_phone": getattr(inv, "bill_to_phone", ""),
            "bill_to_address": inv.bill_to_address or "",
            "customer_tin": getattr(inv, "bill_to_tax_id", ""),
            "items": items,
            # Tax rate percentages for display
            "vat_rate_percent": vat_rate_percent,
            "excise_rate_percent": excise_rate_percent,
        },
        "notes": bilingual_notes,
        # Order grouping for consolidated display
        "order_groups": order_grouping["order_groups"],
        "grouped_grand_total": _money_format(order_grouping["grouped_grand_total"]),
        # Hybrid template flags (for flexible invoice rendering)
        "show_cdf_column": fx_rate_used is not None,
        "show_bank_details": bool(cs.bank_name or cs.mm_provider),
        "exchange_rate": fx_rate_used,
    }

    # Add CDF amounts to line items if exchange rate available
    if fx_rate_used:
        # Add to items (fallback display - these are already dicts)
        for item in context["invoice"]["items"]:
            try:
                # item["line_total"] is a formatted string like "$100.00" - strip currency symbol and commas
                line_total_str = (
                    str(item["line_total"]).replace("$", "").replace(",", "").strip()
                )
                line_total_usd = Decimal(line_total_str)
                item["line_total_cdf"] = (
                    line_total_usd * Decimal(fx_rate_used)
                ).quantize(Decimal("0.01"))
            except Exception:
                item["line_total_cdf"] = None

        # Add to order groups (lines are InvoiceLine objects, not dicts)
        for group in context["order_groups"]:
            # Add CDF attribute to each InvoiceLine object
            for line in group["lines"]:
                try:
                    # InvoiceLine objects have line_total attribute
                    line_total_usd = line.line_total or Decimal("0.00")
                    # Dynamically add CDF attribute to the object
                    line.line_total_cdf = (
                        line_total_usd * Decimal(fx_rate_used)
                    ).quantize(Decimal("0.01"))
                except Exception:
                    line.line_total_cdf = None

    return context


def _build_consolidated_context(cons, cs):
    # Company block reused from single invoice
    address = ", ".join(
        filter(None, [cs.street_address, cs.city, cs.province, cs.country])
    )
    logo_url = None
    try:
        if getattr(cs, "logo", None) and hasattr(cs.logo, "url") and cs.logo.url:
            candidate = cs.logo.url
            resolved = resolve_uri(candidate)
            if resolved and os.path.exists(resolved):
                logo_url = candidate
    except Exception:
        logo_url = None

    company = {
        "legal_name": cs.legal_name or cs.trade_name,
        "trade_name": cs.trade_name or cs.legal_name,
        "email": cs.email,
        "phone": cs.phone,
        "website": cs.website,
        "address": address,
        "street_address": getattr(cs, "street_address", ""),
        "city": getattr(cs, "city", ""),
        "province": getattr(cs, "province", ""),
        "country": getattr(cs, "country", ""),
        "rccm": cs.rccm,
        "id_nat": cs.id_nat,
        "nif": cs.nif,
        "arptc_license": getattr(cs, "arptc_license", ""),
        "tax_regime": getattr(cs, "tax_regime", ""),
        "tax_regime_label": (
            cs.get_tax_regime_display()
            if hasattr(cs, "get_tax_regime_display")
            else getattr(cs, "tax_regime", "")
        ),
        "logo_url": logo_url,
        "currency": getattr(cons, "currency", None) or cs.default_currency or "USD",
        "footer_text": cs.footer_text_en or cs.footer_text_fr,
        "footer_text_fr": getattr(cs, "footer_text_fr", ""),
        "footer_text_en": getattr(cs, "footer_text_en", ""),
        # Banking & Mobile Money
        "bank_name": getattr(cs, "bank_name", ""),
        "bank_account_name": getattr(cs, "bank_account_name", ""),
        "bank_account_number_usd": getattr(cs, "bank_account_number_usd", ""),
        "bank_account_number_cdf": getattr(cs, "bank_account_number_cdf", ""),
        "bank_swift": getattr(cs, "bank_swift", ""),
        "bank_branch": getattr(cs, "bank_branch", ""),
        "bank_iban": getattr(cs, "bank_iban", ""),
        "mm_provider": getattr(cs, "mm_provider", ""),
        "mm_number": getattr(cs, "mm_number", ""),
        "payment_instructions": getattr(cs, "payment_instructions", ""),
        # Branding & Signature
        "signatory_name": getattr(cs, "signatory_name", ""),
        "signatory_title": getattr(cs, "signatory_title", ""),
        "stamp": getattr(cs, "stamp", None),
        "signature": getattr(cs, "signature", None),
        # Compliance & Legal
        "tax_office_name": getattr(cs, "tax_office_name", ""),
        "legal_notes": getattr(cs, "legal_notes", ""),
        # Payment terms
        "payment_terms_days": getattr(cs, "payment_terms_days", 0),
    }

    # Aggregates
    agg_subtotal = Decimal("0.00")
    agg_vat = Decimal("0.00")
    agg_excise = Decimal("0.00")
    agg_tax_total = Decimal("0.00")
    agg_grand = Decimal("0.00")

    child_blocks = []
    for idx, inv in enumerate(cons.child_invoices.all().order_by("issued_at"), start=1):
        subtotal = Decimal(inv.subtotal or 0)
        stored_vat = Decimal(getattr(inv, "vat_amount", 0) or 0)
        stored_exc = Decimal(getattr(inv, "excise_amount", 0) or 0)
        vat_guess = Decimal("0.00")
        exc_guess = Decimal("0.00")

        # Try to infer linked orders to split taxes more accurately
        orders = []
        try:
            orders = [ol.order for ol in inv.order_links.select_related("order")]
            if not orders:
                ords = []
                for ln in inv.lines.all():
                    if getattr(ln, "order", None) and ln.order not in ords:
                        ords.append(ln.order)
                orders = ords
        except Exception:
            orders = []

        if (stored_vat + stored_exc) > 0 or Decimal(inv.tax_total or 0) == 0:
            vat_guess = stored_vat
            exc_guess = stored_exc
            if orders and exc_guess == 0:
                exc_from_orders = Decimal("0.00")
                for o in orders:
                    taxes_qs = o.taxes.all()
                    exc_from_orders += sum(
                        (
                            Decimal(t.amount)
                            for t in taxes_qs
                            if getattr(t, "kind", "").upper() in ("EXCISE", "EXC")
                        ),
                        Decimal("0.00"),
                    )
                if exc_from_orders > 0:
                    exc_guess = exc_from_orders
        else:
            if orders:
                for o in orders:
                    taxes_qs = o.taxes.all()
                    vat_guess += sum(
                        (
                            Decimal(t.amount)
                            for t in taxes_qs
                            if getattr(t, "kind", "").upper() == "VAT"
                        ),
                        Decimal("0.00"),
                    )
                    exc_guess += sum(
                        (
                            Decimal(t.amount)
                            for t in taxes_qs
                            if getattr(t, "kind", "").upper() in ("EXCISE", "EXC")
                        ),
                        Decimal("0.00"),
                    )
            else:
                try:
                    vat_rate = Decimal(
                        inv.vat_rate_percent or cs.vat_rate_percent or 0
                    ) / Decimal("100")
                except Exception:
                    vat_rate = Decimal("0")
                try:
                    exc_rate = (
                        Decimal(inv.excise_rate_percent or cs.excise_rate_percent or 0)
                        / Decimal("100")
                        if (inv.excise_rate_percent or cs.excise_rate_percent)
                        else Decimal("0")
                    )
                except Exception:
                    exc_rate = Decimal("0")
                vat_guess = (
                    (subtotal * vat_rate).quantize(Decimal("0.01"))
                    if vat_rate
                    else Decimal("0.00")
                )
                exc_guess = (
                    (subtotal * exc_rate).quantize(Decimal("0.01"))
                    if exc_rate
                    else Decimal("0.00")
                )

        tax_total = Decimal(inv.tax_total or (vat_guess + exc_guess))
        grand_total = Decimal(inv.grand_total or (subtotal + tax_total))

        # Aggregate
        agg_subtotal += subtotal
        agg_vat += vat_guess
        agg_excise += exc_guess
        agg_tax_total += tax_total
        agg_grand += grand_total

        items = []
        for i, ln in enumerate(
            inv.lines.exclude(kind__iexact="tax")
            .exclude(kind__iexact="tax_adjust")
            .order_by("id"),
            start=1,
        ):
            desc = ln.description or ln.kind or "Item"
            low = (desc or "").lower()
            is_order_ref = False
            try:
                # Match common patterns for order reference markers
                # 1) English/French textual markers (existing)
                if (
                    low.startswith("order ref")
                    or low.startswith("order reference")
                    or "order ref" in low
                ):
                    is_order_ref = True
                elif ("réf" in low or "ref" in low) and (
                    "command" in low or "commande" in low
                ):
                    is_order_ref = True
                elif low.startswith("référence commande") or low.startswith(
                    "reference commande"
                ):
                    is_order_ref = True
                # 2) Explicit order code lines like: "Order ORD-XXXX..."
                #    We consider it a sub‑rubric when the description begins with the word "Order"
                #    (case-insensitive) and contains an "ORD-" code token.
                elif re.search(
                    r"^\s*order\b[^\n]*\bord-[A-Z0-9-]+", desc or "", re.IGNORECASE
                ):
                    is_order_ref = True
                # 3) Kind hint
                elif getattr(ln, "kind", "").lower() in (
                    "order_ref",
                    "order-reference",
                ):
                    is_order_ref = True
            except Exception:
                is_order_ref = False
            items.append(
                {
                    "idx": i,
                    "description": desc,
                    "quantity": ln.quantity,
                    "unit_price": _money_format(ln.unit_price),
                    "line_total": _money_format(ln.line_total),
                    "is_order_ref": is_order_ref,
                }
            )

        # Group invoice lines by order for this child invoice
        order_grouping = group_invoice_lines_by_order(inv)

        child_blocks.append(
            {
                "title": f"Invoice {inv.number}",
                "number": inv.number,
                "issued_at": timezone.localtime(inv.issued_at).strftime("%d %b %Y")
                if inv.issued_at
                else "—",
                "subtotal": _money_format(subtotal),
                "vat": _money_format(vat_guess),
                "excise": _money_format(exc_guess),
                "tax_total": _money_format(tax_total),
                "grand_total": _money_format(grand_total),
                "items": items,
                # Order grouping data
                "order_groups": order_grouping["order_groups"],
                "grouped_grand_total": _money_format(
                    order_grouping["grouped_grand_total"]
                ),
            }
        )

    customer_name = getattr(cons.user, "full_name", None) or getattr(
        cons.user, "username", ""
    )

    # FX equivalent (CDF) for consolidated
    fx_cdf_amount = None
    fx_rate_used = None
    try:
        if (
            getattr(cs, "show_prices_in_cdf", False)
            or (company["currency"] or "USD") == "USD"
        ) and getattr(cons, "issued_at", None):
            fx_rate_used = FxRate.get_rate(cons.issued_at.date(), pair="USD/CDF")
            if fx_rate_used:
                fx_cdf_amount = (agg_grand * Decimal(fx_rate_used)).quantize(
                    Decimal("0.01")
                )
    except Exception:
        fx_rate_used = None
        fx_cdf_amount = None

    # Bilingual notes similar to single invoice
    bilingual_notes = [
        "Prices in USD unless otherwise stated. Official exchange rate applied when displaying CDF. / Prix en USD sauf indication contraire. Taux de change officiel appliqué pour l’affichage en CDF.",
        "This is a consolidated invoice covering multiple invoices listed above. / Ceci est une facture consolidée couvrant plusieurs factures listées ci‑dessus.",
    ]
    if fx_rate_used:
        bilingual_notes.insert(
            0,
            f"FX rate used (USD→CDF): {fx_rate_used} / Taux appliqué (USD→CDF) : {fx_rate_used}",
        )

    # Get tax rate percentages for display (use CompanySettings which may have been populated from TaxRate)
    # For consolidated invoices, we use CompanySettings as the source of truth since
    # individual invoices may have different rates snapshotted
    vat_rate_percent = None
    excise_rate_percent = None
    try:
        vat_rate_percent = Decimal(cs.vat_rate_percent or 0)
    except Exception:
        vat_rate_percent = Decimal("0")
    try:
        excise_rate_percent = (
            Decimal(cs.excise_rate_percent or 0)
            if cs.excise_rate_percent
            else Decimal("0")
        )
    except Exception:
        excise_rate_percent = Decimal("0")

    context = {
        "company": company,
        "consolidated": {
            "number": cons.number,
            "issued_at": (
                timezone.localtime(cons.issued_at).strftime("%d %b %Y")
                if getattr(cons, "issued_at", None)
                else "—"
            ),
            "status": getattr(cons, "status", "").upper(),
            "currency": company["currency"],
            "customer_name": customer_name,
            "subtotal": _money_format(agg_subtotal),
            "vat": _money_format(agg_vat),
            "excise": _money_format(agg_excise),
            "tax_total": _money_format(agg_tax_total),
            "total": _money_format(agg_grand),
            "cdf_amount": (
                f"CDF {fx_cdf_amount:,.2f}" if fx_cdf_amount is not None else None
            ),
            "children": child_blocks,
            # Tax rate percentages for display
            "vat_rate_percent": vat_rate_percent,
            "excise_rate_percent": excise_rate_percent,
        },
        "notes": bilingual_notes,
    }
    return context


# Create your views here.
@login_required
@require_staff_role(["admin", "manager", "finance"])
def billing_management(request):
    template_name = "billing_management_admin.html"
    return render(request, template_name)


@login_required
@require_staff_role(["admin", "finance"])
def set_fx_rate(request):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "message": "Method not allowed"}, status=405
        )
    date_str = (request.POST.get("date") or "").strip()
    pair = (request.POST.get("pair") or "USD/CDF").strip().upper()
    rate_str = (request.POST.get("rate") or "").strip()
    if not date_str or not rate_str:
        return JsonResponse(
            {"success": False, "message": "Missing date or rate"}, status=400
        )
    try:
        when = datetime.fromisoformat(date_str).date()
    except Exception:
        return JsonResponse(
            {"success": False, "message": "Invalid date format"}, status=400
        )
    try:
        rate = Decimal(rate_str)
        if rate <= 0:
            raise ValueError
    except Exception:
        return JsonResponse({"success": False, "message": "Invalid rate"}, status=400)

    # Upsert
    obj, created = FxRate.objects.update_or_create(
        date=when,
        pair=pair,
        defaults={"rate": rate, "created_by": getattr(request, "user", None)},
    )
    msg = "Rate saved successfully." if created else "Rate updated successfully."
    return JsonResponse(
        {
            "success": True,
            "message": msg,
            "date": str(when),
            "pair": pair,
            "rate": str(rate),
        }
    )


@login_required
@require_staff_role(["admin", "manager", "finance"])
def general_ledger(request):
    """
    Return ALL ledger entries (newest -> oldest) across all users.

    Query params:
      - limit: int (default 10, max 200)
      - cursor: opaque string returned as next_cursor
      - include_total: '1' to also return total row count
      - q: free-text (description, user name/email)
      - type: one of entry types (invoice, payment, credit_note, adjustment, tax)
      - account_id: filter by BillingAccount id
      - user_id: filter by custom user PK (id_user)
    """
    # --- pagination inputs ---
    try:
        limit = int(request.GET.get("limit", 10))  # default 10 per page
    except ValueError:
        limit = 10
    limit = max(1, min(limit, 200))

    cursor = request.GET.get("cursor")
    cur_time, cur_id = _decode_cursor(cursor) if cursor else (None, None)

    # --- base queryset (newest first, strict tiebreak on id) ---
    qs = (
        AccountEntry.objects.select_related(
            "account", "account__user", "order", "subscription", "payment"
        )
        .only(
            "id",
            "created_at",
            "entry_type",
            "description",
            "amount_usd",
            "account_id",
            "order_id",
            "subscription_id",
            "payment_id",
            "order__order_reference",
            "account__user__id_user",
            "account__user__full_name",
            "account__user__username",
            "account__user__email",
        )
        .order_by("-created_at", "-id")
    )

    # --- optional filters ---
    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(description__icontains=q)
            | Q(account__user__full_name__icontains=q)
            | Q(account__user__email__icontains=q)
            | Q(order__order_reference__icontains=q)
        )

    typ = (request.GET.get("type") or "").strip()
    if typ:
        qs = qs.filter(entry_type=typ)

    account_id = request.GET.get("account_id")
    if account_id:
        qs = qs.filter(account_id=account_id)

    user_id = request.GET.get("user_id")
    if user_id:
        qs = qs.filter(account__user__id_user=user_id)

    # --- keyset WHERE for "before cursor" ---
    if cur_time and cur_id:
        qs = qs.filter(
            Q(created_at__lt=cur_time) | Q(created_at=cur_time, id__lt=cur_id)
        )

    # --- fetch one extra to detect next page ---
    rows = list(qs[: limit + 1])
    has_more = len(rows) > limit
    if has_more:
        rows = rows[:limit]

    next_cursor = None
    if has_more and rows:
        last = rows[-1]
        next_cursor = _encode_cursor(last.created_at, last.id)

    total = None
    if request.GET.get("include_total") == "1":
        total = qs.model.objects.all().count()  # global count by design

    # --- serialize ---
    data = []
    for e in rows:
        u = getattr(e.account, "user", None)
        # created_at is timezone-aware when USE_TZ=True
        data.append(
            {
                "id": e.id,
                "created_at": e.created_at.isoformat(),
                "entry_type": e.entry_type,
                "entry_type_label": e.get_entry_type_display(),
                "description": e.description or "",
                "amount_usd": str(e.amount_usd),
                # owner context
                "account_id": e.account_id,
                "user_id": getattr(u, "id_user", None),  # custom PK
                "user_name": (
                    getattr(u, "full_name", None) or getattr(u, "username", None)
                ),
                "user_email": getattr(u, "email", None),
                # business links
                "order_reference": getattr(e.order, "order_reference", None),
                "order_id": e.order_id,
                "subscription_id": e.subscription_id,
                "payment_id": e.payment_id,
            }
        )

    payload = {
        "success": True,
        "ledger": data,
        "has_more": has_more,
        "next_cursor": next_cursor,
    }
    if total is not None:
        payload["total"] = total

    return JsonResponse(payload)


# ---------- INVOICE-CENTRIC ENDPOINTS ----------


def _find_invoice_by_identifier(invoice_id: str) -> Invoice:
    ident = (invoice_id or "").strip()
    if not ident:
        raise Http404("Invoice not found")
    ident = unquote(ident)
    # Try by number first
    inv = (
        Invoice.objects.filter(number=ident)
        .prefetch_related("lines", "order_links")
        .select_related("user")
        .first()
    )
    if inv:
        return inv
    # Try integer PK
    try:
        pk = int(ident)
        inv = (
            Invoice.objects.filter(pk=pk)
            .prefetch_related("lines", "order_links")
            .select_related("user")
            .first()
        )
        if inv:
            return inv
    except (TypeError, ValueError):
        pass
    # Fallback: allow searching by alternative fields if present
    inv = (
        Invoice.objects.filter(number__iexact=ident)
        .prefetch_related("lines", "order_links")
        .select_related("user")
        .first()
    )
    if inv:
        return inv
    raise Http404("Invoice not found")


@login_required
def invoice_json_by_number(request, invoice_id: str):
    inv = _find_invoice_by_identifier(invoice_id)
    cs = CompanySettings.get()

    # Prefer stored separated amounts when present; otherwise compute from linked orders
    subtotal = Decimal(inv.subtotal or 0)
    stored_vat = Decimal(getattr(inv, "vat_amount", 0) or 0)
    stored_exc = Decimal(getattr(inv, "excise_amount", 0) or 0)
    vat_guess = Decimal("0.00")
    exc_guess = Decimal("0.00")

    orders = []
    try:
        # Prefer explicit InvoiceOrder links
        orders = [ol.order for ol in inv.order_links.select_related("order")]
        if not orders:
            # Fallback: infer orders from invoice lines
            ords = []
            for ln in inv.lines.all():
                if getattr(ln, "order", None) and ln.order not in ords:
                    ords.append(ln.order)
            orders = ords
    except Exception:
        orders = []

    print("Inside LOOP")

    # Prefer stored if totals were snapshotted, but correct stored zeros when orders clearly have taxes
    if (stored_vat + stored_exc) > 0 or Decimal(inv.tax_total or 0) == 0:
        vat_guess = stored_vat
        exc_guess = stored_exc
        # If Excise is zero in storage but orders indicate excise > 0, override from orders
        if orders and exc_guess == 0:
            exc_from_orders = Decimal("0.00")
            for o in orders:
                taxes_qs = o.taxes.all()
                exc_from_orders += sum(
                    (
                        Decimal(t.amount)
                        for t in taxes_qs
                        if getattr(t, "kind", "").upper() in ("EXCISE", "EXC")
                    ),
                    Decimal("0.00"),
                )
            if exc_from_orders > 0:
                exc_guess = exc_from_orders
    else:
        if orders:
            for o in orders:
                taxes_qs = o.taxes.all()
                vat_guess += sum(
                    (
                        Decimal(t.amount)
                        for t in taxes_qs
                        if getattr(t, "kind", "").upper() == "VAT"
                    ),
                    Decimal("0.00"),
                )
                exc_guess += sum(
                    (
                        Decimal(t.amount)
                        for t in taxes_qs
                        if getattr(t, "kind", "").upper() in ("EXCISE", "EXC")
                    ),
                    Decimal("0.00"),
                )
        else:
            # No orders linked; fallback to snapshot rates
            try:
                vat_rate = Decimal(
                    inv.vat_rate_percent or cs.vat_rate_percent or 0
                ) / Decimal("100")
            except Exception:
                vat_rate = Decimal("0")
            try:
                exc_rate = (
                    Decimal(inv.excise_rate_percent or cs.excise_rate_percent or 0)
                    / Decimal("100")
                    if (inv.excise_rate_percent or cs.excise_rate_percent)
                    else Decimal("0")
                )
            except Exception:
                exc_rate = Decimal("0")
            vat_guess = (
                (subtotal * vat_rate).quantize(Decimal("0.01"))
                if vat_rate
                else Decimal("0.00")
            )
            exc_guess = (
                (subtotal * exc_rate).quantize(Decimal("0.01"))
                if exc_rate
                else Decimal("0.00")
            )

    tax_total = Decimal(inv.tax_total or (vat_guess + exc_guess))

    # FX CDF equivalent for JSON consumers
    fx_rate_used = None
    cdf_equiv = None
    try:
        if inv.issued_at:
            fx_rate_used = FxRate.get_rate(inv.issued_at.date(), pair="USD/CDF")
            if fx_rate_used:
                gt = Decimal(inv.grand_total or (subtotal + tax_total))
                cdf_equiv = (gt * Decimal(fx_rate_used)).quantize(Decimal("0.01"))
    except Exception:
        fx_rate_used = None
        cdf_equiv = None

    items = []
    for ln in inv.lines.all().order_by("id"):
        items.append(
            {
                "description": ln.description,
                "quantity": str(ln.quantity),
                "unit_price": str(ln.unit_price),
                "line_total": str(ln.line_total),
                "kind": ln.kind,
            }
        )

    company = {
        "legal_name": cs.legal_name or cs.trade_name,
        "trade_name": cs.trade_name,
        "email": cs.email,
        "phone": cs.phone,
        "website": cs.website,
        "address": ", ".join(
            filter(None, [cs.street_address, cs.city, cs.province, cs.country])
        ),
        "street_address": getattr(cs, "street_address", ""),
        "city": getattr(cs, "city", ""),
        "province": getattr(cs, "province", ""),
        "country": getattr(cs, "country", ""),
        "rccm": cs.rccm,
        "id_nat": cs.id_nat,
        "nif": cs.nif,
        "tax_regime": cs.get_tax_regime_display()
        if hasattr(cs, "get_tax_regime_display")
        else cs.tax_regime,
        "vat_rate_percent": str(cs.vat_rate_percent),
        "excise_rate_percent": str(cs.excise_rate_percent)
        if cs.excise_rate_percent is not None
        else None,
        "currency": cs.default_currency,
        # Bank details for external consumers
        "bank_name": getattr(cs, "bank_name", ""),
        "bank_account_name": getattr(cs, "bank_account_name", ""),
        "bank_account_number_usd": getattr(cs, "bank_account_number_usd", ""),
        "bank_account_number_cdf": getattr(cs, "bank_account_number_cdf", ""),
        "bank_swift": getattr(cs, "bank_swift", ""),
        "bank_branch": getattr(cs, "bank_branch", ""),
        "bank_iban": getattr(cs, "bank_iban", ""),
    }

    data = {
        "success": True,
        "invoice": {
            "id": inv.id,
            "number": inv.number,
            "status": inv.status,
            "currency": inv.currency,
            "issued_at": inv.issued_at.isoformat() if inv.issued_at else None,
            "due_at": inv.due_at.isoformat() if getattr(inv, "due_at", None) else None,
            "subtotal": str(subtotal),
            "vat": str(vat_guess),
            "excise": str(exc_guess),
            "tax_total": str(tax_total),
            "grand_total": str(inv.grand_total or (subtotal + tax_total)),
            "cdf_equivalent": (str(cdf_equiv) if cdf_equiv is not None else None),
            "fx_rate_used": (str(fx_rate_used) if fx_rate_used is not None else None),
            "bill_to_name": inv.bill_to_name
            or (inv.user.full_name or inv.user.username),
            "bill_to_address": inv.bill_to_address,
            "items": items,
        },
        "company": company,
    }
    return JsonResponse(data)


@login_required
def invoice_pdf_by_number(request, invoice_id: str):
    inv = _find_invoice_by_identifier(invoice_id)
    cs = CompanySettings.get()

    context = _build_invoice_context(inv, cs)

    html = render_to_string(
        "invoices/inv_templates.html", context=context, request=request
    )

    filename = f"{inv.number or 'invoice'}.pdf"
    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'

    # Ensure logo absolute URL if relative
    # With xhtml2pdf, keep media/static URLs relative so resolve_uri can map them
    buf = BytesIO()
    pisa_status = pisa.CreatePDF(
        src=html, dest=buf, link_callback=resolve_uri, encoding="utf-8"
    )
    if getattr(pisa_status, "err", 0):
        log_txt = getattr(pisa_status, "log", None)
        msg = "PDF generation failed for invoice. Please check the template/CSS."
        if log_txt:
            try:
                msg += "\n\n" + str(log_txt)
            except Exception:
                pass
        return HttpResponse(msg, status=500, content_type="text/plain")
    resp.write(buf.getvalue())
    return resp


@login_required
def consolidated_invoice_pdf(request, consolidated_number: str):
    number = unquote((consolidated_number or "").strip())
    if not number:
        raise Http404("Consolidated invoice not found")

    cons = get_object_or_404(
        ConsolidatedInvoice.objects.select_related("user").prefetch_related(
            "child_invoices__lines"
        ),
        number=number,
    )
    cs = CompanySettings.get()

    context = _build_consolidated_context(cons, cs)

    html = render_to_string(
        "invoices/consolidated_inv_templates.html", context=context, request=request
    )

    filename = f"{cons.number or 'consolidated'}.pdf"
    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'

    buf = BytesIO()
    pisa_status = pisa.CreatePDF(
        src=html, dest=buf, link_callback=resolve_uri, encoding="utf-8"
    )
    if getattr(pisa_status, "err", 0):
        log_txt = getattr(pisa_status, "log", None)
        msg = "PDF generation failed for consolidated invoice. Please check the template/CSS."
        if log_txt:
            try:
                msg += "\n\n" + str(log_txt)
            except Exception:
                pass
        return HttpResponse(msg, status=500, content_type="text/plain")
    resp.write(buf.getvalue())
    return resp


# ─────────────────────────────────────────────────────────────────────────────
# Customer Ledger: search + export (XLSX/PDF)
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def ledger_search_customers(request):
    q = (request.GET.get("q") or "").strip()
    if not q:
        return JsonResponse({"success": True, "results": []})
    # Basic search over name/username/email
    users = (
        User.objects.filter(
            Q(full_name__icontains=q) | Q(username__icontains=q) | Q(email__icontains=q)
        )
        .only("id_user", "full_name", "username", "email")
        .order_by("full_name")[:10]
    )
    results = []
    for u in users:
        results.append(
            {
                "id_user": getattr(u, "id_user", None),
                "name": getattr(u, "full_name", None) or getattr(u, "username", ""),
                "email": getattr(u, "email", ""),
            }
        )
    return JsonResponse({"success": True, "results": results})


def _parse_iso_date(s: str):
    try:
        if not s:
            return None
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


@login_required
def ledger_export(request):
    # Params
    user_id = (request.GET.get("user_id") or "").strip()
    if not user_id:
        return JsonResponse(
            {"success": False, "message": "Missing user_id"}, status=400
        )
    try:
        uid = int(user_id)
    except ValueError:
        return JsonResponse(
            {"success": False, "message": "Invalid user_id"}, status=400
        )

    date_from = _parse_iso_date((request.GET.get("from") or "").strip())
    date_to = _parse_iso_date((request.GET.get("to") or "").strip())
    typ = (request.GET.get("type") or "").strip()
    fmt = (request.GET.get("format") or "xlsx").strip().lower()
    include_cdf = (request.GET.get("include_cdf") or "1").strip() in (
        "1",
        "true",
        "yes",
    )

    # Base queryset: filter by user primary key on BillingAccount relation
    entries = (
        AccountEntry.objects.select_related(
            "account__user", "order", "subscription", "payment"
        )
        .filter(account__user__id_user=uid)
        .order_by("created_at", "id")
    )
    if typ:
        entries = entries.filter(entry_type=typ)
    if date_from:
        entries = entries.filter(created_at__date__gte=date_from)
    if date_to:
        entries = entries.filter(created_at__date__lte=date_to)

    # Safeguard: limit excessive exports without date range
    max_rows = 10000
    rows_count = entries.count()
    if not date_from and not date_to and rows_count > max_rows:
        return JsonResponse(
            {
                "success": False,
                "message": f"Too many rows ({rows_count}). Please provide a date range.",
            },
            status=400,
        )

    # Opening balance if from-date provided
    opening_balance = Decimal("0.00")
    if date_from:
        before_qs = AccountEntry.objects.filter(account__user__id_user=uid)
        if typ:
            before_qs = before_qs.filter(entry_type=typ)
        opening_balance = before_qs.filter(created_at__date__lt=date_from).aggregate(
            total=models.Sum("amount_usd")
        )["total"] or Decimal("0.00")

    # Build rows and compute running balance and CDF
    rows = []
    run_usd = opening_balance
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")
    total_cdf = Decimal("0.00")

    def _debit_credit(amount: Decimal):
        amt = Decimal(amount or 0)
        if amt >= 0:
            return amt, Decimal("0.00")
        else:
            return Decimal("0.00"), -amt

    for e in entries.iterator():
        amt = Decimal(e.amount_usd or 0)
        debit, credit = _debit_credit(amt)
        run_usd += amt
        # CDF per-row
        cdf_amt = None
        cdf_run = None
        if include_cdf:
            try:
                rate = FxRate.get_rate(e.created_at.date(), pair="USD/CDF")
                if rate:
                    cdf_amt = (amt * Decimal(rate)).quantize(Decimal("0.01"))
                    cdf_run = (run_usd * Decimal(rate)).quantize(Decimal("0.01"))
            except Exception:
                cdf_amt = None
                cdf_run = None
        total_debit += debit
        total_credit += credit
        if cdf_amt is not None:
            total_cdf += cdf_amt
        rows.append(
            {
                "date": e.created_at.strftime("%Y-%m-%d"),
                "type": e.get_entry_type_display(),
                "description": e.description or "",
                "debit_usd": debit,
                "credit_usd": credit,
                "balance_usd": run_usd,
                "amount_cdf": cdf_amt,
                "balance_cdf": cdf_run,
                "order_ref": getattr(e.order, "order_reference", "") or "",
                "subscription_id": e.subscription_id,
                "payment_id": e.payment_id,
            }
        )

    # Dispatch by format
    if fmt == "pdf":
        return _ledger_pdf_response(
            uid, date_from, date_to, rows, opening_balance, include_cdf
        )
    else:
        return _ledger_xlsx_response(
            uid,
            date_from,
            date_to,
            rows,
            opening_balance,
            include_cdf,
            total_debit,
            total_credit,
            total_cdf,
        )


def _ledger_xlsx_response(
    uid,
    date_from,
    date_to,
    rows,
    opening_balance,
    include_cdf,
    total_debit,
    total_credit,
    total_cdf,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"

    # Headers (bilingual)
    headers = [
        "Date",
        "Type · Type",
        "Description · Description",
        "Debit (USD) · Débit (USD)",
        "Credit (USD) · Crédit (USD)",
        "Balance (USD) · Solde (USD)",
    ]
    if include_cdf:
        headers += [
            "Amount (CDF) · Montant (CDF)",
            "Balance (CDF) · Solde (CDF)",
        ]
    headers += [
        "Order Ref · Réf. commande",
        "Subscription ID",
        "Payment Ref",
    ]
    ws.append(headers)

    # Opening balance row
    ob_row = [
        date_from.strftime("%Y-%m-%d") if date_from else "",
        "Opening Balance · Solde d'ouverture",
        "",
        "",
        "",
        float(opening_balance),
    ]
    if include_cdf:
        ob_row += ["", ""]
    ob_row += ["", "", ""]
    ws.append(ob_row)

    # Data rows
    for r in rows:
        row = [
            r["date"],
            r["type"],
            r["description"],
            float(r["debit_usd"]),
            float(r["credit_usd"]),
            float(r["balance_usd"]),
        ]
        if include_cdf:
            row += [
                float(r["amount_cdf"]) if r.get("amount_cdf") is not None else "",
                float(r["balance_cdf"]) if r.get("balance_cdf") is not None else "",
            ]
        row += [
            r["order_ref"],
            r["subscription_id"] or "",
            r["payment_id"] or "",
        ]
        ws.append(row)

    # Totals row
    totals_label = "Totals · Totaux"
    totals_row = ["", totals_label, "", float(total_debit), float(total_credit), ""]
    if include_cdf:
        totals_row += [float(total_cdf), ""]
    totals_row += ["", "", ""]
    ws.append(totals_row)

    # Basic styling: freeze header, autofilter, column widths
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(45, max(12, max_len + 2))

    # HTTP response
    filename = f"ledger_{uid}_{(date_from or '')}_{(date_to or '')}.xlsx"
    from django.utils.encoding import smart_str

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f"attachment; filename={smart_str(filename)}"
    from io import BytesIO as _BIO

    bio = _BIO()
    wb.save(bio)
    resp.write(bio.getvalue())
    return resp


def _ledger_pdf_response(uid, date_from, date_to, rows, opening_balance, include_cdf):
    cs = CompanySettings.get()
    context = {
        "company": cs,
        "filters": {
            "user_id": uid,
            "from": date_from.strftime("%Y-%m-%d") if date_from else None,
            "to": date_to.strftime("%Y-%m-%d") if date_to else None,
            "include_cdf": include_cdf,
        },
        "opening_balance": opening_balance,
        "rows": rows,
    }
    html = render_to_string("ledgers/customer_ledger.html", context=context)
    filename = f"ledger_{uid}_{(date_from or '')}_{(date_to or '')}.pdf"
    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    buf = BytesIO()
    pisa_status = pisa.CreatePDF(
        src=html, dest=buf, link_callback=resolve_uri, encoding="utf-8"
    )
    if getattr(pisa_status, "err", 0):
        return HttpResponse(
            "PDF generation failed for ledger.", status=500, content_type="text/plain"
        )
    resp.write(buf.getvalue())
    return resp


@login_required
def ledger_statement(request):
    # Params
    user_id = (request.GET.get("user_id") or "").strip()
    if not user_id:
        return JsonResponse(
            {"success": False, "message": "Missing user_id"}, status=400
        )
    try:
        uid = int(user_id)
    except ValueError:
        return JsonResponse(
            {"success": False, "message": "Invalid user_id"}, status=400
        )

    date_from = _parse_iso_date((request.GET.get("from") or "").strip())
    date_to = _parse_iso_date((request.GET.get("to") or "").strip())
    fmt = (request.GET.get("format") or "xlsx").strip().lower()
    include_cdf = (request.GET.get("include_cdf") or "1").strip() in (
        "1",
        "true",
        "yes",
    )

    # Default date_to = today if not provided
    if not date_to:
        try:
            date_to = timezone.localdate()
        except Exception:
            date_to = datetime.now(dt_tz.utc).date()

    # Opening balance strictly before from
    opening_balance = Decimal("0.00")
    if date_from:
        before_qs = AccountEntry.objects.filter(account__user__id_user=uid)
        opening_balance = before_qs.filter(created_at__date__lt=date_from).aggregate(
            total=models.Sum("amount_usd")
        )["total"] or Decimal("0.00")

    # Period transactions [from, to]
    entries = (
        AccountEntry.objects.select_related(
            "account__user", "order", "subscription", "payment"
        )
        .filter(account__user__id_user=uid)
        .order_by("created_at", "id")
    )
    if date_from:
        entries = entries.filter(created_at__date__gte=date_from)
    if date_to:
        entries = entries.filter(created_at__date__lte=date_to)

    rows = []
    run_usd = opening_balance
    total_debit = Decimal("0.00")
    total_credit = Decimal("0.00")

    def _debit_credit(amount: Decimal):
        amt = Decimal(amount or 0)
        if amt >= 0:
            return amt, Decimal("0.00")
        else:
            return Decimal("0.00"), -amt

    for e in entries.iterator():
        amt = Decimal(e.amount_usd or 0)
        debit, credit = _debit_credit(amt)
        run_usd += amt
        row = {
            "date": e.created_at.strftime("%Y-%m-%d"),
            "type": e.get_entry_type_display(),
            "description": e.description or "",
            "debit_usd": debit,
            "credit_usd": credit,
            "balance_usd": run_usd,
            "order_ref": getattr(e.order, "order_reference", "") or "",
            "subscription_id": e.subscription_id,
            "payment_id": e.payment_id,
        }
        rows.append(row)
        total_debit += debit
        total_credit += credit

    closing_balance = run_usd

    # Aging buckets based on invoices due dates as of date_to
    aging = {
        "d0_30": Decimal("0.00"),
        "d31_60": Decimal("0.00"),
        "d61_90": Decimal("0.00"),
        "d90_plus": Decimal("0.00"),
    }
    inv_qs = (
        Invoice.objects.filter(user__id_user=uid)
        .exclude(status__in=[Invoice.Status.PAID, Invoice.Status.CANCELLED])
        .only("due_at", "grand_total", "issued_at", "status")
    )
    for inv in inv_qs.iterator():
        amt = Decimal(getattr(inv, "grand_total", 0) or 0)
        due = getattr(inv, "due_at", None)
        if not due:
            continue
        due_date = (
            timezone.localtime(due).date() if timezone.is_aware(due) else due.date()
        )
        days_past = (date_to - due_date).days
        if days_past <= 0:
            continue  # not due yet
        if days_past <= 30:
            aging["0_30"] += amt
        elif days_past <= 60:
            aging["31_60"] += amt
        elif days_past <= 90:
            aging["61_90"] += amt
        else:
            aging["90_plus"] += amt

    # Optional CDF equivalents (as of date_to)
    cdf_rate = None
    cdf = None
    try:
        cdf_rate = FxRate.get_rate(date_to, pair="USD/CDF") if include_cdf else None
    except Exception:
        cdf_rate = None
    if cdf_rate:

        def conv(x):
            try:
                return (Decimal(x) * Decimal(cdf_rate)).quantize(Decimal("0.01"))
            except Exception:
                return None

        cdf = {
            "opening": conv(opening_balance),
            "closing": conv(closing_balance),
            "total_debit": conv(total_debit),
            "total_credit": conv(total_credit),
            "aging": {k: conv(v) for k, v in aging.items()},
            "rate": str(cdf_rate),
        }

    if fmt == "json":
        return JsonResponse(
            {
                "success": True,
                "statement": {
                    "user_id": uid,
                    "from": date_from.strftime("%Y-%m-%d") if date_from else None,
                    "to": date_to.strftime("%Y-%m-%d") if date_to else None,
                    "opening_balance": str(opening_balance),
                    "total_debit": str(total_debit),
                    "total_credit": str(total_credit),
                    "closing_balance": str(closing_balance),
                    "aging": {k: str(v) for k, v in aging.items()},
                    "cdf": cdf,
                    "rows": [
                        {
                            **{
                                k: (str(v) if isinstance(v, Decimal) else v)
                                for k, v in r.items()
                                if k in ("debit_usd", "credit_usd", "balance_usd")
                            },
                            **{
                                k: v
                                for k, v in r.items()
                                if k not in ("debit_usd", "credit_usd", "balance_usd")
                            },
                        }
                        for r in rows
                    ],
                },
            }
        )
    elif fmt == "pdf":
        return _statement_pdf_response(
            uid,
            date_from,
            date_to,
            opening_balance,
            total_debit,
            total_credit,
            closing_balance,
            aging,
            cdf,
        )
    else:
        return _statement_xlsx_response(
            uid,
            date_from,
            date_to,
            opening_balance,
            total_debit,
            total_credit,
            closing_balance,
            aging,
            cdf,
        )


def _statement_pdf_response(
    uid, date_from, date_to, opening, total_debit, total_credit, closing, aging, cdf
):
    cs = CompanySettings.get()
    context = {
        "company": cs,
        "filters": {
            "user_id": uid,
            "from": date_from.strftime("%Y-%m-%d") if date_from else None,
            "to": date_to.strftime("%Y-%m-%d") if date_to else None,
        },
        "opening": opening,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "closing": closing,
        "aging": aging,
        "cdf": cdf,
    }
    html = render_to_string("ledgers/customer_statement.html", context=context)
    filename = f"statement_{uid}_{(date_from or '')}_{(date_to or '')}.pdf"
    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    buf = BytesIO()
    pisa_status = pisa.CreatePDF(
        src=html, dest=buf, link_callback=resolve_uri, encoding="utf-8"
    )
    if getattr(pisa_status, "err", 0):
        return HttpResponse(
            "PDF generation failed for statement.",
            status=500,
            content_type="text/plain",
        )
    resp.write(buf.getvalue())
    return resp


def _statement_xlsx_response(
    uid, date_from, date_to, opening, total_debit, total_credit, closing, aging, cdf
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"

    headers = [
        "Customer Statement · Relevé client",
        f"User ID: {uid}",
        f"From: {date_from.strftime('%Y-%m-%d') if date_from else ''}",
        f"To: {date_to.strftime('%Y-%m-%d') if date_to else ''}",
    ]
    ws.append(headers)
    ws.append(
        [
            "Opening · Ouverture",
            float(opening),
            "Total Debit · Débit",
            float(total_debit),
            "Total Credit · Crédit",
            float(total_credit),
            "Closing · Clôture",
            float(closing),
        ]
    )

    ws.append([])
    ws.append(["Aging · Échéancier", "0–30", "31–60", "61–90", "90+"])
    ws.append(
        [
            "USD",
            float(aging.get("0_30", 0)),
            float(aging.get("31_60", 0)),
            float(aging.get("61_90", 0)),
            float(aging.get("90_plus", 0)),
        ]
    )
    if cdf:
        ws.append(
            [
                "CDF",
                float(cdf["aging"].get("0_30", 0) or 0),
                float(cdf["aging"].get("31_60", 0) or 0),
                float(cdf["aging"].get("61_90", 0) or 0),
                float(cdf["aging"].get("90_plus", 0) or 0),
            ]
        )

    # Column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    # HTTP response
    filename = f"statement_{uid}_{(date_from or '')}_{(date_to or '')}.xlsx"
    from django.utils.encoding import smart_str

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f"attachment; filename={smart_str(filename)}"
    from io import BytesIO as _BIO

    bio = _BIO()
    wb.save(bio)
    resp.write(bio.getvalue())
    return resp


# Revenue reporting endpoints


@login_required
@require_staff_role(["admin", "manager", "finance"])
def revenue_summary(request):
    """
    Summary aggregates for revenue reporting.

    Query params:
      - from: YYYY-MM-DD
      - to:   YYYY-MM-DD
      - perspective: invoiced|collected (default: invoiced)
      - region: optional region name/id filter (by Order.region.name exact or id)
      - include_cdf: 1|0 (default 1)
    """
    date_from = _parse_iso_date((request.GET.get("from") or "").strip())
    date_to = _parse_iso_date((request.GET.get("to") or "").strip())
    perspective = (request.GET.get("perspective") or "invoiced").strip().lower()
    include_cdf = (request.GET.get("include_cdf") or "1").strip() in (
        "1",
        "true",
        "yes",
    )
    region_filter = (request.GET.get("region") or "").strip()

    if not date_from or not date_to:
        return JsonResponse(
            {"success": False, "message": "Missing date range (from/to)"}, status=400
        )
    if (date_to - date_from).days < 0:
        return JsonResponse(
            {"success": False, "message": "Invalid date range"}, status=400
        )
    # Cap span to 366 days unless explicitly allowed (guardrail)
    if (date_to - date_from).days > 366:
        return JsonResponse(
            {
                "success": False,
                "message": "Date range too large; please use 366 days or less.",
            },
            status=400,
        )

    # Helpers
    def _region_ok(order_region):
        if not region_filter:
            return True
        if order_region is None:
            return False
        try:
            # Allow numeric id or name match
            if region_filter.isdigit():
                return str(getattr(order_region, "id", "")) == region_filter
            # name exact or case-insensitive
            return (
                getattr(order_region, "name", "") or ""
            ).strip().lower() == region_filter.lower()
        except Exception:
            return False

    # Aggregates containers
    totals_usd = {
        "subtotal": Decimal("0.00"),
        "vat": Decimal("0.00"),
        "excise": Decimal("0.00"),
        "tax_total": Decimal("0.00"),
        "grand_total": Decimal("0.00"),
        "collected": Decimal("0.00"),
        "credits_adjustments": Decimal("0.00"),
        "net": Decimal("0.00"),
    }
    totals_cdf = None
    if include_cdf:
        totals_cdf = {
            k: Decimal("0.00")
            for k in [
                "subtotal",
                "vat",
                "excise",
                "tax_total",
                "grand_total",
                "collected",
                "credits_adjustments",
                "net",
            ]
        }

    if perspective == "invoiced":
        inv_qs = (
            Invoice.objects.select_related("user")
            .prefetch_related("order_links__order__region")
            .filter(issued_at__date__gte=date_from, issued_at__date__lte=date_to)
            .exclude(status=Invoice.Status.CANCELLED)
        )
        for inv in inv_qs.iterator(chunk_size=1000):
            # Region filter using any linked order's region; if no links, allow Unknown only when no region filter supplied
            regions = []
            try:
                regions = [
                    ol.order.region
                    for ol in inv.order_links.all()
                    if getattr(ol, "order", None)
                ]
            except Exception:
                regions = []
            if region_filter:
                if not regions:
                    continue
                if not any(_region_ok(r) for r in regions if r is not None):
                    continue

            subtotal = Decimal(inv.subtotal or 0)
            vat = Decimal(getattr(inv, "vat_amount", 0) or 0)
            exc = Decimal(getattr(inv, "excise_amount", 0) or 0)
            tax_total = Decimal(inv.tax_total or (vat + exc))
            grand_total = Decimal(inv.grand_total or (subtotal + tax_total))

            totals_usd["subtotal"] += subtotal
            totals_usd["vat"] += vat
            totals_usd["excise"] += exc
            totals_usd["tax_total"] += tax_total
            totals_usd["grand_total"] += grand_total

            if include_cdf and inv.issued_at:
                rate = FxRate.get_rate(inv.issued_at.date(), pair="USD/CDF")
                if rate:
                    r = Decimal(rate)
                    totals_cdf["subtotal"] += (subtotal * r).quantize(Decimal("0.01"))
                    totals_cdf["vat"] += (vat * r).quantize(Decimal("0.01"))
                    totals_cdf["excise"] += (exc * r).quantize(Decimal("0.01"))
                    totals_cdf["tax_total"] += (tax_total * r).quantize(Decimal("0.01"))
                    totals_cdf["grand_total"] += (grand_total * r).quantize(
                        Decimal("0.01")
                    )

        # For summary, net = grand_total - credits_adjustments (if provided separately); we compute credits separately below
        # Credits/Adjustments within window (deductions)
        entries = AccountEntry.objects.select_related("order__region").filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
            entry_type__in=("credit_note", "adjustment"),
        )
        if region_filter:
            entries = entries.filter(order__region__isnull=False)
        for e in entries.iterator():
            if region_filter and not _region_ok(getattr(e.order, "region", None)):
                continue
            amt = Decimal(e.amount_usd or 0)
            totals_usd["credits_adjustments"] += (
                -amt if amt < 0 else amt
            )  # store as positive deduction magnitude
            if include_cdf:
                rate = FxRate.get_rate(e.created_at.date(), pair="USD/CDF")
                if rate:
                    totals_cdf["credits_adjustments"] += (
                        abs(amt) * Decimal(rate)
                    ).quantize(Decimal("0.01"))

        totals_usd["net"] = (
            totals_usd["grand_total"] - totals_usd["credits_adjustments"]
        )
        if include_cdf and totals_cdf is not None:
            totals_cdf["net"] = (
                totals_cdf["grand_total"] - totals_cdf["credits_adjustments"]
            )

    else:  # collected
        pay_qs = AccountEntry.objects.select_related("order__region").filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
            entry_type="payment",
        )
        for e in pay_qs.iterator():
            if region_filter and not _region_ok(getattr(e.order, "region", None)):
                continue
            amt = Decimal(e.amount_usd or 0)
            totals_usd["collected"] += amt if amt > 0 else Decimal("0.00")
            if include_cdf:
                rate = FxRate.get_rate(e.created_at.date(), pair="USD/CDF")
                if rate:
                    totals_cdf["collected"] += (
                        max(amt, Decimal("0.00")) * Decimal(rate)
                    ).quantize(Decimal("0.01"))
        # Credits/Adjustments in collected perspective shown as separate
        cred_qs = AccountEntry.objects.select_related("order__region").filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
            entry_type__in=("credit_note", "adjustment"),
        )
        for e in cred_qs.iterator():
            if region_filter and not _region_ok(getattr(e.order, "region", None)):
                continue
            amt = Decimal(e.amount_usd or 0)
            # positive magnitude for presentation
            totals_usd["credits_adjustments"] += -amt if amt < 0 else amt
            if include_cdf:
                rate = FxRate.get_rate(e.created_at.date(), pair="USD/CDF")
                if rate:
                    totals_cdf["credits_adjustments"] += (
                        abs(amt) * Decimal(rate)
                    ).quantize(Decimal("0.01"))

    payload = {
        "success": True,
        "perspective": perspective,
        "from": str(date_from),
        "to": str(date_to),
        "totals_usd": {
            k: str(v.quantize(Decimal("0.01"))) for k, v in totals_usd.items()
        },
    }
    if include_cdf and totals_cdf is not None:
        payload["totals_cdf"] = {
            k: str(v.quantize(Decimal("0.01"))) for k, v in totals_cdf.items()
        }
    # Backwards compatibility: older callers expect a `totals` key with USD totals.
    # Provide `totals` as an alias to `totals_usd` so existing clients/tests don't break.
    if "totals_usd" in payload:
        payload["totals"] = payload["totals_usd"]
    return JsonResponse(payload)


@login_required
@require_staff_role(["admin", "manager", "finance"])
def revenue_table(request):
    """
    Grouped revenue table.

    Query params:
      - from, to (required)
      - group_by: day|week|month|region (default: month)
      - perspective: invoiced|collected (default: invoiced)
      - include_cdf: 1|0 (default 1)
      - page, size (optional, default size 100)
    """
    date_from = _parse_iso_date((request.GET.get("from") or "").strip())
    date_to = _parse_iso_date((request.GET.get("to") or "").strip())
    group_by = (request.GET.get("group_by") or "month").strip().lower()
    perspective = (request.GET.get("perspective") or "invoiced").strip().lower()
    include_cdf = (request.GET.get("include_cdf") or "1").strip() in (
        "1",
        "true",
        "yes",
    )

    try:
        page = max(1, int(request.GET.get("page", 1)))
    except Exception:
        page = 1
    try:
        size = max(1, min(200, int(request.GET.get("size", 100))))
    except Exception:
        size = 100

    if group_by not in ("day", "week", "month", "region"):
        return JsonResponse(
            {"success": False, "message": "Invalid group_by"}, status=400
        )
    if perspective not in ("invoiced", "collected"):
        return JsonResponse(
            {"success": False, "message": "Invalid perspective"}, status=400
        )
    if not date_from or not date_to:
        return JsonResponse(
            {"success": False, "message": "Missing date range (from/to)"}, status=400
        )
    if (date_to - date_from).days > 366:
        return JsonResponse(
            {
                "success": False,
                "message": "Date range too large; please use 366 days or less.",
            },
            status=400,
        )

    # event iterator per perspective
    events = []  # list of dicts: {date, region_label, metrics: {...}, cdf: {...}}

    def _region_label_from_order(order):
        try:
            r = getattr(order, "region", None)
            if not r:
                return "Unknown"
            name = getattr(r, "name", None)
            return (name or "Unknown").strip() or "Unknown"
        except Exception:
            return "Unknown"

    if perspective == "invoiced":
        inv_qs = (
            Invoice.objects.select_related("user")
            .prefetch_related("order_links__order__region")
            .filter(issued_at__date__gte=date_from, issued_at__date__lte=date_to)
            .exclude(status=Invoice.Status.CANCELLED)
            .order_by("issued_at", "id")
        )
        for inv in inv_qs.iterator(chunk_size=1000):
            subtotal = Decimal(inv.subtotal or 0)
            vat = Decimal(getattr(inv, "vat_amount", 0) or 0)
            exc = Decimal(getattr(inv, "excise_amount", 0) or 0)
            tax_total = Decimal(inv.tax_total or (vat + exc))
            grand_total = Decimal(inv.grand_total or (subtotal + tax_total))
            d = inv.issued_at.date() if inv.issued_at else None
            regions = []
            try:
                regions = [
                    ol.order
                    for ol in inv.order_links.all()
                    if getattr(ol, "order", None)
                ]
            except Exception:
                regions = []
            region_label = (
                _region_label_from_order(regions[0]) if regions else "Unknown"
            )

            evt = {
                "date": d,
                "region": region_label,
                "usd": {
                    "subtotal": subtotal,
                    "vat": vat,
                    "excise": exc,
                    "tax_total": tax_total,
                    "grand_total": grand_total,
                    "credits_adjustments": Decimal("0.00"),
                    "collected": Decimal("0.00"),
                    "net": grand_total,  # will be adjusted if we apply credits at group level
                },
                "cdf": None,
            }
            if include_cdf and d:
                rate = FxRate.get_rate(d, pair="USD/CDF")
                if rate:
                    r = Decimal(rate)
                    evt["cdf"] = {
                        "subtotal": (subtotal * r).quantize(Decimal("0.01")),
                        "vat": (vat * r).quantize(Decimal("0.01")),
                        "excise": (exc * r).quantize(Decimal("0.01")),
                        "tax_total": (tax_total * r).quantize(Decimal("0.01")),
                        "grand_total": (grand_total * r).quantize(Decimal("0.01")),
                        "credits_adjustments": Decimal("0.00"),
                        "collected": Decimal("0.00"),
                        "net": (grand_total * r).quantize(Decimal("0.01")),
                    }
            events.append(evt)

        # Credits/Adjustments: attach as separate events (negative net) for grouping math
        cred_qs = (
            AccountEntry.objects.select_related("order__region")
            .filter(
                created_at__date__gte=date_from,
                created_at__date__lte=date_to,
                entry_type__in=("credit_note", "adjustment"),
            )
            .order_by("created_at", "id")
        )
        for e in cred_qs.iterator():
            amt = Decimal(e.amount_usd or 0)
            mag = -amt if amt < 0 else amt
            d = e.created_at.date() if e.created_at else None
            reg_label = _region_label_from_order(getattr(e, "order", None))
            evt = {
                "date": d,
                "region": reg_label,
                "usd": {
                    "subtotal": Decimal("0.00"),
                    "vat": Decimal("0.00"),
                    "excise": Decimal("0.00"),
                    "tax_total": Decimal("0.00"),
                    "grand_total": Decimal("0.00"),
                    "credits_adjustments": mag,
                    "collected": Decimal("0.00"),
                    "net": -mag,  # subtract from net
                },
                "cdf": None,
            }
            if include_cdf and d:
                rate = FxRate.get_rate(d, pair="USD/CDF")
                if rate:
                    r = Decimal(rate)
                    evt["cdf"] = {
                        "subtotal": Decimal("0.00"),
                        "vat": Decimal("0.00"),
                        "excise": Decimal("0.00"),
                        "tax_total": Decimal("0.00"),
                        "grand_total": Decimal("0.00"),
                        "credits_adjustments": (mag * r).quantize(Decimal("0.01")),
                        "collected": Decimal("0.00"),
                        "net": (-(mag) * r).quantize(Decimal("0.01")),
                    }
            events.append(evt)

    else:  # collected
        pay_qs = (
            AccountEntry.objects.select_related("order__region")
            .filter(
                created_at__date__gte=date_from,
                created_at__date__lte=date_to,
                entry_type="payment",
            )
            .order_by("created_at", "id")
        )
        for e in pay_qs.iterator():
            amt = Decimal(e.amount_usd or 0)
            if amt <= 0:
                continue
            d = e.created_at.date() if e.created_at else None
            reg_label = _region_label_from_order(getattr(e, "order", None))
            evt = {
                "date": d,
                "region": reg_label,
                "usd": {
                    "subtotal": Decimal("0.00"),
                    "vat": Decimal("0.00"),
                    "excise": Decimal("0.00"),
                    "tax_total": Decimal("0.00"),
                    "grand_total": Decimal("0.00"),
                    "credits_adjustments": Decimal("0.00"),
                    "collected": amt,
                    "net": amt,
                },
                "cdf": None,
            }
            if include_cdf and d:
                rate = FxRate.get_rate(d, pair="USD/CDF")
                if rate:
                    r = Decimal(rate)
                    evt["cdf"] = {
                        "subtotal": Decimal("0.00"),
                        "vat": Decimal("0.00"),
                        "excise": Decimal("0.00"),
                        "tax_total": Decimal("0.00"),
                        "grand_total": Decimal("0.00"),
                        "credits_adjustments": Decimal("0.00"),
                        "collected": (amt * r).quantize(Decimal("0.01")),
                        "net": (amt * r).quantize(Decimal("0.01")),
                    }
            events.append(evt)
        # Credits/Adjustments as separate negative events for collected perspective too
        cred_qs = (
            AccountEntry.objects.select_related("order__region")
            .filter(
                created_at__date__gte=date_from,
                created_at__date__lte=date_to,
                entry_type__in=("credit_note", "adjustment"),
            )
            .order_by("created_at", "id")
        )
        for e in cred_qs.iterator():
            amt = Decimal(e.amount_usd or 0)
            mag = -amt if amt < 0 else amt
            d = e.created_at.date() if e.created_at else None
            reg_label = _region_label_from_order(getattr(e, "order", None))
            evt = {
                "date": d,
                "region": reg_label,
                "usd": {
                    "subtotal": Decimal("0.00"),
                    "vat": Decimal("0.00"),
                    "excise": Decimal("0.00"),
                    "tax_total": Decimal("0.00"),
                    "grand_total": Decimal("0.00"),
                    "credits_adjustments": mag,
                    "collected": Decimal("0.00"),
                    "net": -mag,
                },
                "cdf": None,
            }
            if include_cdf and d:
                rate = FxRate.get_rate(d, pair="USD/CDF")
                if rate:
                    r = Decimal(rate)
                    evt["cdf"] = {
                        "subtotal": Decimal("0.00"),
                        "vat": Decimal("0.00"),
                        "excise": Decimal("0.00"),
                        "tax_total": Decimal("0.00"),
                        "grand_total": Decimal("0.00"),
                        "credits_adjustments": (mag * r).quantize(Decimal("0.01")),
                        "collected": Decimal("0.00"),
                        "net": (-(mag) * r).quantize(Decimal("0.01")),
                    }
            events.append(evt)

    # Group
    from collections import defaultdict

    groups = defaultdict(
        lambda: {
            "usd": {
                k: Decimal("0.00")
                for k in [
                    "subtotal",
                    "vat",
                    "excise",
                    "tax_total",
                    "grand_total",
                    "credits_adjustments",
                    "collected",
                    "net",
                ]
            },
            "cdf": (
                {
                    k: Decimal("0.00")
                    for k in [
                        "subtotal",
                        "vat",
                        "excise",
                        "tax_total",
                        "grand_total",
                        "credits_adjustments",
                        "collected",
                        "net",
                    ]
                }
                if include_cdf
                else None
            ),
            "label": None,
        }
    )

    def _key_and_label(evt):
        d = evt.get("date")
        reg = evt.get("region")
        if group_by == "day":
            return (d.isoformat() if d else "Unknown"), (
                d.isoformat() if d else "Unknown"
            )
        if group_by == "week":
            if d:
                iso_year, iso_week, _ = d.isocalendar()
                label = f"{iso_year}-W{iso_week:02d}"
            else:
                label = "Unknown"
            return label, label
        if group_by == "month":
            if d:
                label = f"{d.year}-{d.month:02d}"
            else:
                label = "Unknown"
            return label, label
        # region
        label = reg or "Unknown"
        return label, label

    for evt in events:
        key, label = _key_and_label(evt)
        grp = groups[key]
        grp["label"] = label
        for k, v in evt["usd"].items():
            grp["usd"][k] += Decimal(v or 0)
        if include_cdf and evt.get("cdf") is not None:
            for k, v in evt["cdf"].items():
                grp["cdf"][k] += Decimal(v or 0)

    # Order groups by key asc
    ordered = sorted(groups.items(), key=lambda kv: kv[0])

    # Pagination (simple page/size over grouped rows)
    total_groups = len(ordered)
    start = (page - 1) * size
    end = min(start + size, total_groups)
    page_items = ordered[start:end]

    def _fmt(d):
        return {k: str(Decimal(v).quantize(Decimal("0.01"))) for k, v in d.items()}

    rows = []
    for gk, gv in page_items:
        row = {
            "key": gk,
            "label": gv["label"],
            "usd": _fmt(gv["usd"]),
        }
        if include_cdf and gv.get("cdf") is not None:
            row["cdf"] = _fmt(gv["cdf"])
        rows.append(row)

    payload = {
        "success": True,
        "group_by": group_by,
        "perspective": perspective,
        "from": str(date_from),
        "to": str(date_to),
        "rows": rows,
        "page": page,
        "size": size,
        "total_groups": total_groups,
        "has_more": end < total_groups,
    }
    return JsonResponse(payload)


# ─────────────────────────────────────────────────────────────────────────────
# Region reconciliation endpoints: check and fix a single Order.region
# ─────────────────────────────────────────────────────────────────────────────
@login_required
@require_GET
def check_order_region(request):
    """Compare the Order.region saved value against the region resolved from GPS.

    Query params:
      - order_id: required (int primary key of Order)

    Returns JSON with saved vs resolved region and diagnostics.
    """
    order_id = (request.GET.get("order_id") or "").strip()
    if not order_id:
        return JsonResponse(
            {"success": False, "message": "Missing order_id"}, status=400
        )
    try:
        oid = int(order_id)
    except ValueError:
        return JsonResponse(
            {"success": False, "message": "Invalid order_id"}, status=400
        )

    order = (
        Order.objects.select_related("region")
        .only("id", "region_id", "latitude", "longitude")
        .filter(pk=oid)
        .first()
    )
    if not order:
        return JsonResponse(
            {"success": False, "message": "Order not found"}, status=404
        )

    resolved_region, tag = resolve_region_from_coords(order.latitude, order.longitude)

    def region_payload(r):
        if not r:
            return None
        return {"id": getattr(r, "id", None), "name": getattr(r, "name", None)}

    saved = getattr(order, "region", None)
    match = (
        saved is not None
        and resolved_region is not None
        and getattr(saved, "id", None) == getattr(resolved_region, "id", None)
    )

    data = {
        "success": True,
        "order_id": order.id,
        "coords": {"lat": order.latitude, "lon": order.longitude},
        "saved_region": region_payload(saved),
        "resolved_region": region_payload(resolved_region),
        "match": bool(match),
        "resolver_tag": tag,  # auto | auto_ambiguous | no_coords | no_match
    }
    # Helpful boolean flags
    data["has_coords"] = order.latitude is not None and order.longitude is not None
    data["ambiguous"] = tag == "auto_ambiguous"
    data["no_match"] = tag == "no_match"
    return JsonResponse(data)


@login_required
@require_POST
def fix_order_region(request):
    """Attempt to set `Order.region` based on GPS resolution.

    Body (form-encoded or JSON):
      - order_id: required
      - force: optional (1/true) — allow applying even if ambiguous; otherwise ambiguous cases are rejected.
      - override: optional (1/true) — allow overriding an existing saved region when it differs.
    """
    order_id = (request.POST.get("order_id") or "").strip()
    if not order_id:
        return JsonResponse(
            {"success": False, "message": "Missing order_id"}, status=400
        )
    try:
        oid = int(order_id)
    except ValueError:
        return JsonResponse(
            {"success": False, "message": "Invalid order_id"}, status=400
        )

    # Permissions: staff/admin only to apply fixes via API
    user = getattr(request, "user", None)
    if not getattr(user, "is_staff", False):
        return JsonResponse({"success": False, "message": "Forbidden"}, status=403)

    force = (request.POST.get("force") or "").strip().lower() in {"1", "true", "yes"}
    override = (request.POST.get("override") or "").strip().lower() in {
        "1",
        "true",
        "yes",
    }

    order = Order.objects.select_related("region").filter(pk=oid).first()
    if not order:
        return JsonResponse(
            {"success": False, "message": "Order not found"}, status=404
        )

    resolved_region, tag = resolve_region_from_coords(order.latitude, order.longitude)

    if tag in ("no_coords", "no_match"):
        return JsonResponse(
            {"success": False, "message": f"Cannot resolve region: {tag}"}, status=409
        )
    if tag == "auto_ambiguous" and not force:
        return JsonResponse(
            {
                "success": False,
                "message": "Ambiguous match; pass force=1 to override explicitly.",
            },
            status=409,
        )

    if not resolved_region:
        return JsonResponse(
            {"success": False, "message": "No region resolved."}, status=409
        )

    # If there is a saved region and it's different, require override unless none is set
    if (
        getattr(order, "region_id", None)
        and getattr(order, "region_id", None) != getattr(resolved_region, "id", None)
        and not override
    ):
        return JsonResponse(
            {
                "success": False,
                "message": "Saved region differs; pass override=1 to apply.",
            },
            status=409,
        )

    # Apply update
    order.region_id = getattr(resolved_region, "id", None)
    order.save(update_fields=["region"])

    return JsonResponse(
        {
            "success": True,
            "message": "Region updated from GPS.",
            "order_id": order.id,
            "saved_region": {
                "id": getattr(order.region, "id", None),
                "name": getattr(order.region, "name", None),
            },
            "resolver_tag": tag,
        }
    )
