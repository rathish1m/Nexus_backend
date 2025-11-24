import json
import random
import string
import time
from io import BytesIO

import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST

from client_app.views import settings
from geo_regions.models import Region
from main.models import (
    StarlinkKit,
    StarlinkKitInventory,
    StarlinkKitMovement,
    StockLocation,
)
from user.permissions import require_staff_role


# Create your views here.
@login_required(login_url="login_page")
@require_staff_role(["admin", "finance"])
def stock_management(request):
    template = "stock_management_page.html"
    return render(request, template)


@login_required(login_url="login_page")
@require_staff_role(["admin", "finance"])
def download_stock_sample(request):
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "New Starlink Stock"

    # Define headers for new stock upload
    headers = [
        "kit_number",  # e.g., STLK-9087234
        "serial_number",  # e.g., SN-123456789
        "model",  # e.g., Dishy v2
        "firmware_version",  # e.g., v1.0.5
        "kit_id",  # ForeignKey to StarlinkKit
        "location",  # Location of StarlinkKit
    ]

    # Add header row
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font

    # Add sample row
    ws.append(
        [
            "STLK-9087234",
            "SN-123456789",
            "Dishy v2",
            "v1.0.5",
            1,  # example StarlinkKit ID
            False,
        ]
    )

    # Save workbook to memory
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # Return file as response
    response = HttpResponse(
        file_stream,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        'attachment; filename="starlink_stock_sample.xlsx"'
    )
    return response


@login_required(login_url="login_page")
@require_staff_role(["admin", "finance"])
@require_POST
def upload_stock_excel(request):
    excel_file = request.FILES.get("file")
    if not excel_file:
        return JsonResponse(
            {"success": False, "message": "No file provided."}, status=400
        )

    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Failed to read Excel: {e}"}, status=400
        )

    headers = [cell.value for cell in ws[1]]
    expected_headers = [
        "kit_number",
        "serial_number",
        "model",
        "firmware_version",
        "kit_id",
        "location",
    ]

    if headers != expected_headers:
        return JsonResponse(
            {
                "success": False,
                "message": "Invalid Excel format. Headers must be: "
                + ", ".join(expected_headers),
            },
            status=400,
        )

    inventory_records = []
    movement_records = []
    created_kit_numbers = []

    # Preprocess Excel rows
    excel_rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in excel_rows:
        row_values = list(row)[:6]  # Trim to max 6 columns
        if len(row_values) < 6:
            continue  # Skip incomplete rows

        kit_number, serial_number, model, firmware_version, kit_id, location = (
            row_values
        )

        if (
            not kit_number
            or StarlinkKitInventory.objects.filter(kit_number=kit_number).exists()
        ):
            continue  # Skip duplicates or missing kit_number

        try:
            kit = StarlinkKit.objects.get(id=kit_id)
        except StarlinkKit.DoesNotExist:
            continue  # Skip if kit type is invalid

        # Prepare inventory item
        inventory_item = StarlinkKitInventory(
            kit_number=kit_number,
            serial_number=serial_number,
            model=model,
            firmware_version=firmware_version,
            kit=kit,
            is_assigned=False,
        )
        inventory_records.append(inventory_item)
        created_kit_numbers.append(kit_number)

    # Bulk insert inventory
    StarlinkKitInventory.objects.bulk_create(inventory_records)

    # Fetch saved inventory with IDs
    new_inventories = StarlinkKitInventory.objects.filter(
        kit_number__in=created_kit_numbers
    )

    # Match locations again by kit_number
    for item in new_inventories:
        # Find the Excel row with matching kit_number
        row_match = next((r for r in excel_rows if r[0] == item.kit_number), None)
        location = row_match[5] if row_match and len(row_match) > 5 else None

        movement = StarlinkKitMovement(
            inventory_item=item,
            movement_type="received",
            timestamp=timezone.now(),
            location=location,
            note="Bulk received via Excel upload",
            created_by=request.user,
        )
        movement_records.append(movement)

    # Bulk insert movement records
    StarlinkKitMovement.objects.bulk_create(movement_records)

    return JsonResponse(
        {
            "success": True,
            "message": f"{len(inventory_records)} new kits added and movements recorded.",
        }
    )


# ---------- Helpers ----------


def _ok(payload=None, **extra):
    base = {"success": True}
    if payload:
        base.update(payload)
    base.update(extra)
    return JsonResponse(base)


def _err(message="Error", status=400, **extra):
    data = {"success": False, "message": message}
    data.update(extra)
    return JsonResponse(data, status=status)


# ---------- Regions ----------


# ----------------- helpers -----------------
def _parse_json(request):
    try:
        return json.loads(request.body.decode("utf-8"))
    except Exception:
        return None


def _gen_kit_number(kit_id: int) -> str:
    """Simple unique-looking kit number for bulk 'stock_add' creations."""
    ts = int(time.time())
    suffix = "".join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"K{kit_id}-{ts}-{suffix}"


# ----------------- regions -----------------
@require_GET
@login_required
def get_regions(request):
    """
    Return available regions from geo_regions.Region.
    We try to return id, name, code when present (code may not exist on your Region model).
    """
    fields = ["id", "name"]
    # include code only if field exists on the model
    if any(f.name == "code" for f in Region._meta.fields):
        regions = list(Region.objects.order_by("name").values("id", "name", "code"))
    else:
        regions = list(Region.objects.order_by("name").values("id", "name"))
    return JsonResponse({"success": True, "regions": regions})


@require_GET
@login_required
def get_configured_region(request):
    """
    If settings.STOCK_DEFAULT_REGION_ID is set, return it.
    Else if only one exists, return that (helps lock the dropdown in your UI).
    """
    region_dict = None
    rid = getattr(settings, "STOCK_DEFAULT_REGION_ID", None)
    if rid:
        region = Region.objects.filter(id=rid).first()
        if region:
            region_dict = {
                "id": region.id,
                "name": getattr(region, "name", ""),
                "code": getattr(region, "code", None),
            }
    if not region_dict and Region.objects.count() == 1:
        r = Region.objects.first()
        region_dict = {
            "id": r.id,
            "name": getattr(r, "name", ""),
            "code": getattr(r, "code", None),
        }

    if not region_dict:
        return JsonResponse({"success": False, "message": "No configured region."})
    return JsonResponse({"success": True, "region": region_dict})


# ----------------- locations -----------------
@require_GET
@login_required
def get_stock_locations(request):
    """
    Optional filter: ?region_id=<id>
    """
    region_id = request.GET.get("region_id")
    qs = StockLocation.objects.select_related("region").order_by("name")
    if region_id:
        qs = qs.filter(region_id=region_id)
    locations = [
        {
            "id": loc.id,
            "name": loc.name,
            "code": loc.code,
            "address": loc.address,
            "region_id": loc.region_id,
            "region": getattr(loc.region, "name", None) if loc.region_id else None,
            "is_active": loc.is_active,
        }
        for loc in qs
    ]
    return JsonResponse({"success": True, "locations": locations})


@require_POST
@login_required
def create_stock_location(request):
    """
    Body: {code, name, address?, region_id?}
    """
    data = _parse_json(request)
    if not data:
        return JsonResponse({"success": False, "message": "Invalid JSON."}, status=400)

    code = (data.get("code") or "").strip()
    name = (data.get("name") or "").strip()
    address = (data.get("address") or "").strip()
    region_id = data.get("region_id")

    if not code or not name:
        return JsonResponse(
            {"success": False, "message": "Code and Name are required."}, status=400
        )

    region = None
    if region_id:
        region = Region.objects.filter(id=region_id).first()
        if not region:
            return JsonResponse(
                {"success": False, "message": "Region not found."}, status=404
            )

    loc = StockLocation.objects.create(
        code=code, name=name, address=address, region=region
    )
    return JsonResponse({"success": True, "id": loc.id, "message": "Location created."})


# ----------------- kits -----------------
@require_GET
@login_required
def get_kits(request):
    """
    Return active StarlinkKit choices for the “kit” dropdowns.
    """
    kits = list(
        StarlinkKit.objects.filter(is_active=True)
        .order_by("name")
        .values("id", "name", "model", "description", "kit_type", "base_price_usd")
    )
    return JsonResponse({"success": True, "kits": kits})


# ----------------- inventory grid -----------------
@require_GET
@login_required
def stock_with_quantity(request):
    group_by = (request.GET.get("group_by") or "location").strip().lower()
    region_id = request.GET.get("region_id")
    location_id = request.GET.get("location_id")
    kit_type = request.GET.get("kit_type")

    include_assigned = request.GET.get("include_assigned") == "1"
    include_scrapped = request.GET.get("include_scrapped") == "1"

    base = StarlinkKitInventory.objects.select_related(
        "kit", "current_location", "current_location__region"
    ).filter(current_location__isnull=False)

    if region_id:
        base = base.filter(current_location__region_id=region_id)
    if location_id:
        base = base.filter(current_location_id=location_id)
    if kit_type:
        base = base.filter(kit__kit_type=kit_type)

    if group_by == "region":
        values = [
            "kit_id",
            "kit__name",
            "kit__model",
            "kit__description",
            "kit__kit_type",
            "current_location__region_id",
            "current_location__region__name",
        ]
        order = ["kit__name", "current_location__region__name"]
    else:
        values = [
            "kit_id",
            "kit__name",
            "kit__model",
            "kit__description",
            "kit__kit_type",
            "current_location_id",
            "current_location__name",
            "current_location__region_id",
            "current_location__region__name",
        ]
        order = ["kit__name", "current_location__name"]

    qs = (
        base.values(*values)
        .annotate(
            # not scrapped AND not assigned
            available_quantity=Count(
                "id",
                filter=(~Q(status__iexact="scrapped")) & Q(is_assigned=False),
            ),
            # not scrapped AND assigned
            assigned_quantity=Count(
                "id",
                filter=(~Q(status__iexact="scrapped")) & Q(is_assigned=True),
            ),
            # scrapped only
            scrapped_quantity=Count("id", filter=Q(status__iexact="scrapped")),
            # everything not scrapped
            active_quantity=Count("id", filter=~Q(status__iexact="scrapped")),
            # all rows
            total_quantity=Count("id"),
        )
        .order_by(*order)
    )

    rows = []
    for r in qs:
        qty = r["available_quantity"]
        if include_assigned:
            qty += r["assigned_quantity"]
        if include_scrapped:
            qty += r["scrapped_quantity"]

        if group_by == "region":
            rows.append(
                {
                    "id": r["kit_id"],
                    "item_id": r["kit_id"],
                    "kit": r["kit__name"] or "—",
                    "model": r["kit__model"] or "",
                    "kit_type": r["kit__kit_type"] or "",
                    "description": r["kit__description"] or "",
                    "region_id": r["current_location__region_id"],
                    "region": r["current_location__region__name"],
                    "location_id": None,
                    "location": None,
                    "quantity": int(qty),
                    "available_quantity": int(r["available_quantity"] or 0),
                    "assigned_quantity": int(r["assigned_quantity"] or 0),
                    "scrapped_quantity": int(r["scrapped_quantity"] or 0),
                    "active_quantity": int(r["active_quantity"] or 0),
                    "total_quantity": int(r["total_quantity"] or 0),
                }
            )
        else:
            rows.append(
                {
                    "id": r["kit_id"],
                    "item_id": r["kit_id"],
                    "kit": r["kit__name"] or "—",
                    "model": r["kit__model"] or "",
                    "kit_type": r["kit__kit_type"] or "",
                    "description": r["kit__description"] or "",
                    "location_id": r["current_location_id"],
                    "location": r["current_location__name"],
                    "region_id": r["current_location__region_id"],
                    "region": r["current_location__region__name"],
                    "quantity": int(qty),
                    "available_quantity": int(r["available_quantity"] or 0),
                    "assigned_quantity": int(r["assigned_quantity"] or 0),
                    "scrapped_quantity": int(r["scrapped_quantity"] or 0),
                    "active_quantity": int(r["active_quantity"] or 0),
                    "total_quantity": int(r["total_quantity"] or 0),
                }
            )

    return JsonResponse({"success": True, "stocks": rows})


# ----------------- add stock (receive) -----------------
@require_POST
@login_required
@transaction.atomic
def stock_add(request):
    """
    Body: { item_id, location_id, quantity }
    Creates `quantity` StarlinkKitInventory rows of the selected StarlinkKit
    and places them at `location_id`. Also logs movements as 'received'.
    """
    data = _parse_json(request)
    if not data:
        return JsonResponse({"success": False, "message": "Invalid JSON."}, status=400)

    item_id = data.get("item_id")
    location_id = data.get("location_id")
    try:
        qty = int(data.get("quantity"))
    except Exception:
        qty = None

    if not item_id or not location_id or not qty or qty < 1:
        return JsonResponse(
            {
                "success": False,
                "message": "item_id, location_id and positive quantity are required.",
            },
            status=400,
        )

    kit = StarlinkKit.objects.filter(id=item_id, is_active=True).first()
    if not kit:
        return JsonResponse({"success": False, "message": "Kit not found."}, status=404)

    loc = StockLocation.objects.filter(id=location_id, is_active=True).first()
    if not loc:
        return JsonResponse(
            {"success": False, "message": "Location not found."}, status=404
        )

    created_ids = []
    for _ in range(qty):
        inv = StarlinkKitInventory.objects.create(
            kit=kit,
            model=kit.model,
            current_location=loc,
            status="available",
            condition=StarlinkKitInventory.Condition.NEW,
            kit_number=_gen_kit_number(kit.id),
        )
        created_ids.append(inv.id)
        StarlinkKitMovement.objects.create(
            inventory_item=inv,
            movement_type="received",
            location=loc.name,
            note="Stock received",
            created_by=request.user if request.user.is_authenticated else None,
        )

    return JsonResponse(
        {"success": True, "count": len(created_ids), "message": "Stock updated."}
    )


# ----------------- transfer (between locations/regions) -----------------
@require_POST
@login_required
@transaction.atomic
def move_stock_between_regions(request):
    """
    Body: { item_id, from_location_id, to_location_id, quantity }
    Moves `quantity` available items of a given StarlinkKit from one location to another.
    Creates movement logs ('transferred') for each item moved.
    """
    data = _parse_json(request)
    if not data:
        return JsonResponse({"success": False, "message": "Invalid JSON."}, status=400)

    item_id = data.get("item_id")
    from_location_id = data.get("from_location_id")
    to_location_id = data.get("to_location_id")
    try:
        qty = int(data.get("quantity"))
    except Exception:
        qty = None

    if not item_id or not from_location_id or not to_location_id or not qty or qty < 1:
        return JsonResponse(
            {
                "success": False,
                "message": "All fields and positive quantity are required.",
            },
            status=400,
        )
    if from_location_id == to_location_id:
        return JsonResponse(
            {"success": False, "message": "Source and destination cannot be the same."},
            status=400,
        )

    kit = StarlinkKit.objects.filter(id=item_id, is_active=True).first()
    if not kit:
        return JsonResponse({"success": False, "message": "Kit not found."}, status=404)

    from_loc = StockLocation.objects.filter(id=from_location_id, is_active=True).first()
    to_loc = StockLocation.objects.filter(id=to_location_id, is_active=True).first()
    if not from_loc or not to_loc:
        return JsonResponse(
            {"success": False, "message": "Location not found."}, status=404
        )

    # Select available items to move
    qs = (
        StarlinkKitInventory.objects.select_for_update()
        .filter(
            kit_id=kit.id,
            current_location_id=from_loc.id,
            is_assigned=False,
        )
        .exclude(status="scrapped")
        .order_by("id")
    )
    to_move = list(qs[:qty])
    if len(to_move) < qty:
        return JsonResponse(
            {"success": False, "message": "Insufficient stock in source location."},
            status=400,
        )

    for inv in to_move:
        inv.current_location = to_loc
        inv.save(update_fields=["current_location"])
        StarlinkKitMovement.objects.create(
            inventory_item=inv,
            movement_type="transferred",
            location=to_loc.name,
            note=f"Transfer {from_loc.name} → {to_loc.name}",
            created_by=request.user if request.user.is_authenticated else None,
        )

    return JsonResponse(
        {"success": True, "moved": len(to_move), "message": "Stock moved successfully."}
    )


@login_required
@require_POST
def create_region(request):
    """
    Create a Region.

    Body (JSON):
      - name: str (required)
      - code: str (optional, used only if Region has a 'code' field)
      - is_active: bool (optional, used only if Region has 'is_active')
      - parent_id: int (optional FK if Region has 'parent')
      - country_id: int (optional FK if Region has 'country')

    Returns:
      { success, id, region: {...}, message }
    """
    data = _parse_json(request)
    if not data:
        return JsonResponse({"success": False, "message": "Invalid JSON."}, status=400)

    name = (data.get("name") or "").strip()
    if not name:
        return JsonResponse(
            {"success": False, "message": "Field 'name' is required."}, status=400
        )

    # discover available fields on Region dynamically
    field_names = {f.name for f in Region._meta.get_fields()}
    create_kwargs = {"name": name}

    # optional simple fields
    if "code" in field_names and data.get("code") is not None:
        create_kwargs["code"] = (data.get("code") or "").strip()

    if "is_active" in field_names and data.get("is_active") is not None:
        create_kwargs["is_active"] = bool(data.get("is_active"))

    # optional FKs: parent, country
    # We assign via *_id if the field exists to avoid extra queries.
    if "parent" in field_names and data.get("parent_id") is not None:
        create_kwargs["parent_id"] = data.get("parent_id") or None

    if "country" in field_names and data.get("country_id") is not None:
        create_kwargs["country_id"] = data.get("country_id") or None

    try:
        region = Region.objects.create(**create_kwargs)
    except IntegrityError as e:
        # handle common unique/name/code errors gracefully
        return JsonResponse(
            {"success": False, "message": f"Could not create region: {str(e)}"},
            status=400,
        )

    # build a lightweight payload back
    payload = {"id": region.id, "name": getattr(region, "name", None)}
    if "code" in field_names:
        payload["code"] = getattr(region, "code", None)
    if "is_active" in field_names:
        payload["is_active"] = getattr(region, "is_active", True)
    if "parent" in field_names:
        payload["parent_id"] = getattr(region, "parent_id", None)
    if "country" in field_names:
        payload["country_id"] = getattr(region, "country_id", None)

    return JsonResponse(
        {
            "success": True,
            "id": region.id,
            "region": payload,
            "message": "Region created.",
        }
    )


@login_required
@require_GET
def get_kits(request):
    """
    Return Starlink kits for selects.

    Optional query params:
      - q: search text across name/model/description
      - kit_type: "standard" | "mini"
      - is_active: "1"/"true" or "0"/"false" (default true)
      - limit: int (default 100)

    Response:
      {
        "success": true,
        "kits": [
          {
            "id": 1,
            "name": "Starlink Standard",
            "model": "V4",
            "description": "...",
            "kit_type": "standard",
            "base_price_usd": "599.00",
            "picture_url": "https://.../media/..."
          },
          ...
        ],
        "count": 2
      }
    """
    q = (request.GET.get("q") or "").strip()
    kit_type = (request.GET.get("kit_type") or "").strip().lower()
    is_active_param = (request.GET.get("is_active") or "1").strip().lower()
    limit = request.GET.get("limit")

    # Guard limit
    try:
        limit = max(1, min(int(limit), 500)) if limit else 100
    except (TypeError, ValueError):
        limit = 100

    qs = StarlinkKit.objects.all()

    # Default to active kits
    if is_active_param in ("0", "false", "no"):
        qs = qs.filter(is_active=False)
    else:
        qs = qs.filter(is_active=True)

    if kit_type in ("standard", "mini"):
        qs = qs.filter(kit_type=kit_type)

    if q:
        qs = qs.filter(
            Q(name__icontains=q) | Q(model__icontains=q) | Q(description__icontains=q)
        )

    qs = qs.order_by("name")[:limit]

    def kit_to_dict(k: StarlinkKit):
        pic_url = None
        try:
            if k.picture:
                pic_url = k.picture.url
        except Exception:
            pic_url = None

        return {
            "id": k.id,
            "name": k.name,
            "model": k.model or "",
            "description": k.description or "",
            "kit_type": k.kit_type,
            "base_price_usd": str(k.base_price_usd),
            "picture_url": pic_url,
        }

    data = [kit_to_dict(k) for k in qs]
    return JsonResponse({"success": True, "kits": data, "count": len(data)})
